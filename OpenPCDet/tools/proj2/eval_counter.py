import json, pickle
import os
from sklearn.metrics import accuracy_score
import math
import torch
import torch.nn.functional as F
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import Workbook
import numpy as np
from sklearn.metrics import confusion_matrix
import cv2
import random
from datetime import datetime

def load_eval_file(file_path):
    if os.path.exists(file_path):
        result_dict = json.load(open(file_path))
    return result_dict


def sigmoid(x):
    y = torch.clamp(x.sigmoid(), min=1e-4, max=1 - 1e-4)
    return y

def extract_centers(heatmap, obj_name, thresh=0.5, radius=2):
    device = heatmap.device
    batch, c, h, w = heatmap.shape
    all_centers = []

    # print('heat map shape: {}'.format(heatmap.shape))
    # print(heatmap[1, 0, :, :])

    # if obj_name == 'motorcycle' or obj_name == 'bicycle':
    #     thresh = 0.6

    for b in range(batch):
        batch_centers = []
        for ch in range(c):
            heatmap_bch = heatmap[b, ch, :, :]

            # print(heatmap_bch.shape)
            hm_np = heatmap.numpy()
            hm = (hm_np * 255).astype('uint8')

            # Apply threshold to the heatmap
            keep = heatmap_bch > thresh
            heatmap_bch = heatmap_bch * keep.float()

            # Suppress non-local maxima
            heatmap_max = F.max_pool2d(heatmap_bch.unsqueeze(0).unsqueeze(0), (2 * radius + 1, 2 * radius + 1), stride=1, padding=radius)
            keep = (heatmap_bch == heatmap_max.squeeze(0).squeeze(0))
            heatmap_bch = heatmap_bch * keep.float()

            # Get indices of the center points
            y_idx, x_idx = torch.nonzero(heatmap_bch, as_tuple=True)

            for i in range(len(y_idx)):
                batch_centers.append((x_idx[i].item(), y_idx[i].item()))
        
        all_centers.append(batch_centers)

    # Convert list of centers to tensor
    centers_tensor = [torch.tensor(centers, dtype=torch.float32, device=device) for centers in all_centers]
    max_len = max(len(centers) for centers in centers_tensor)
    padded_centers = torch.zeros(batch, max_len, 2, device=device)
    for i, centers in enumerate(centers_tensor):
        if len(centers) > 0:
            padded_centers[i, :len(centers), :] = centers
    
    num_centers = torch.tensor([len(centers) for centers in all_centers], dtype=torch.float32, device=device)

    # print('num_centers: {}'.format(num_centers))
    # print('padded centers: ')
    # print(padded_centers)
    
    return num_centers, padded_centers

def extract_centers_dynamic(heatmap, thresh=0.5, radius=2):
    device = heatmap.device
    batch, c, h, w = heatmap.shape
    all_centers = []

    # print('heat map shape: {}'.format(heatmap.shape))
    # print(heatmap[1, 0, :, :])
    for b in range(batch):
        batch_centers = []
        for ch in range(c):
            heatmap_bch = heatmap[b, ch, :, :]

            # print(heatmap_bch.shape)
            hm_np = heatmap.numpy()
            hm = (hm_np * 255).astype('uint8')
            
            ret, otsu_thresh = cv2.threshold(hm, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU) # ret is the treshold
            # print(ret/255 + 0.1)
            if ret == 0:
                dynamic_thresh = 1
            else:
                dynamic_thresh = ret/255 + thresh
            # print('thresh: {} '.format(dynamic_thresh))

            # Apply threshold to the heatmap
            keep = heatmap_bch > dynamic_thresh
            heatmap_bch = heatmap_bch * keep.float()

            # Suppress non-local maxima
            heatmap_max = F.max_pool2d(heatmap_bch.unsqueeze(0).unsqueeze(0), (2 * radius + 1, 2 * radius + 1), stride=1, padding=radius)
            keep = (heatmap_bch == heatmap_max.squeeze(0).squeeze(0))
            heatmap_bch = heatmap_bch * keep.float()

            # Get indices of the center points
            y_idx, x_idx = torch.nonzero(heatmap_bch, as_tuple=True)

            for i in range(len(y_idx)):
                batch_centers.append((x_idx[i].item(), y_idx[i].item()))
        
        all_centers.append(batch_centers)

    # Convert list of centers to tensor
    centers_tensor = [torch.tensor(centers, dtype=torch.float32, device=device) for centers in all_centers]
    max_len = max(len(centers) for centers in centers_tensor)
    padded_centers = torch.zeros(batch, max_len, 2, device=device)
    for i, centers in enumerate(centers_tensor):
        if len(centers) > 0:
            padded_centers[i, :len(centers), :] = centers
    
    num_centers = torch.tensor([len(centers) for centers in all_centers], dtype=torch.float32, device=device)

    # print('num_centers: {}'.format(num_centers))
    # print('padded centers: ')
    # print(padded_centers)
    
    return num_centers, padded_centers

def extract_centers_dynamic2(heatmap, obj_name, thresh=0.5, radius=2):
    device = heatmap.device
    batch, c, h, w = heatmap.shape
    all_centers = []

    # print('heat map shape: {}'.format(heatmap.shape))
    # print(heatmap[1, 0, :, :])
    for b in range(batch):
        batch_centers = []
        for ch in range(c):
            heatmap_bch = heatmap[b, ch, :, :]

            # print(heatmap_bch.shape)
            hm_np = heatmap.numpy()
            hm = (hm_np * 255).astype('uint8')
            
            ret, otsu_thresh = cv2.threshold(hm, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU) # ret is the treshold
            # print(ret/255 + 0.1)

            if obj_name == 'motorcycle' or obj_name == 'bicycle':
                thresh = 0.12
            if ret == 0:
                dynamic_thresh = 1
            else:
                dynamic_thresh = ret/255 + thresh
            # print('thresh: {} '.format(dynamic_thresh))

            # Apply threshold to the heatmap
            keep = heatmap_bch > dynamic_thresh
            heatmap_bch = heatmap_bch * keep.float()

            # Suppress non-local maxima
            heatmap_max = F.max_pool2d(heatmap_bch.unsqueeze(0).unsqueeze(0), (2 * radius + 1, 2 * radius + 1), stride=1, padding=radius)
            keep = (heatmap_bch == heatmap_max.squeeze(0).squeeze(0))
            heatmap_bch = heatmap_bch * keep.float()

            # Get indices of the center points
            y_idx, x_idx = torch.nonzero(heatmap_bch, as_tuple=True)

            for i in range(len(y_idx)):
                batch_centers.append((x_idx[i].item(), y_idx[i].item()))
        
        all_centers.append(batch_centers)

    # Convert list of centers to tensor
    centers_tensor = [torch.tensor(centers, dtype=torch.float32, device=device) for centers in all_centers]
    max_len = max(len(centers) for centers in centers_tensor)
    padded_centers = torch.zeros(batch, max_len, 2, device=device)
    for i, centers in enumerate(centers_tensor):
        if len(centers) > 0:
            padded_centers[i, :len(centers), :] = centers
    
    num_centers = torch.tensor([len(centers) for centers in all_centers], dtype=torch.float32, device=device)

    # print('num_centers: {}'.format(num_centers))
    # print('padded centers: ')
    # print(padded_centers)
    
    return num_centers, padded_centers


def hm(thresh, radius):
    print('thresh: {}, radius: {}'.format(thresh, radius))
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    # objs = ['car', 'pedestrian', 'cyclist']
    gt_path = 'gt_list.txt'#'/data/kitti_gt_list3.txt'#'/data/kitti_gt_list2.txt' #'gt_list.txt'
    if os.path.exists(gt_path):
        gt_list = json.load(open(gt_path))

    result_dict = {}
    for i in range(15050): #15050
        obj_idx = i % 3
        file_path = '/data/hms/{}.txt'.format(i)
        if os.path.exists(file_path):
            hm= json.load(open(file_path))
            hm= torch.tensor(hm, dtype=torch.float32)
            # hm=hm.unsqueeze(1)
            # hm = hm + 25
            
            # print(hm)
            # hm = hm.permute(1, 0, 2, 3)
            hm = sigmoid(hm)
            # print(hm)
            num_centers, padded_centers = extract_centers(hm, objs[obj_idx], thresh, radius=radius)
            # print('number of centers: {}'.format(num_centers))
            
            num_centers = num_centers.tolist()
            gt = gt_list[i]

        obj = objs[obj_idx]
        if obj in result_dict.keys():
            result_dict[obj]['gt_count'].extend(gt)
            result_dict[obj]['center_count'].extend(num_centers)
        else:
            result_dict[obj] = {'gt_count': gt, 'center_count': num_centers}
        
        # if obj == 'car':
        #     print('gt count: {}, pred: {}'.format(gt, num_centers))

        if i % 2000 == 0:
            print(i)
    # eval2(result_dict)
    file = 'result_output/kitti/result_counternet2_{}_{}.txt'.format(thresh, radius)
    # file = 'result_output/kitti_result_counternet.txt'
    with open(file, 'w') as f: 
        json.dump(result_dict, f)
    return result_dict

def remove_duplicate_centers(tensor, radius=3):
    # Get the batch size and dimensions of each feature map
    batch_size, height, width = tensor.shape

    # Process each batch independently
    for b in range(batch_size):
        for i in range(height):
            for j in range(width):
                # If the current center is 1, process its neighborhood
                if tensor[b, i, j] == 1:
                    # Calculate the neighborhood bounds
                    min_x = max(0, i - radius)
                    max_x = min(height, i + radius + 1)
                    min_y = max(0, j - radius)
                    max_y = min(width, j + radius + 1)
                    
                    # Set the surrounding centers within the radius to 0, except the center itself
                    for x in range(min_x, max_x):
                        for y in range(min_y, max_y):
                            # Check if the point is within the circle of given radius
                            if (x - i)**2 + (y - j)**2 <= radius**2:
                                if x != i or y != j:
                                    tensor[b, x, y] = 0
    return tensor

def hm_partition_multiscale(thresh, radius, radius_duplicate = 2):
    print('thresh: {}, radius: {}, radius duplicate: {}'.format(thresh, radius, radius_duplicate))
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    gt_path = 'gt_list.txt'
    if os.path.exists(gt_path):
            gt_list = json.load(open(gt_path))

    result_dict = {}
    ranges = [[0, 64, 0, 128], [64, 128, 0, 128], [0, 128, 0, 64], [0, 128, 64, 128], [0, 64, 0, 64],[0, 64, 64, 128],[64, 128, 0, 64],[64, 128, 64, 128]]
    for i in range(15050): #15050
        obj_idx = i % 10
        file_path = '/data/hms_partition_multiscale_ep3/{}.txt'.format(i)
        if os.path.exists(file_path):
            hms= json.load(open(file_path))
            # partition number of hms
            maps = np.zeros((4, 128, 128)) # batch size
            scales = [0,0,1,1,2,2,2,2]
            for hm_idx, hm in enumerate(hms):
                scale_id = scales[hm_idx]
                coord_range = ranges[hm_idx]
                hm= torch.tensor(hm, dtype=torch.float32)
                hm = sigmoid(hm)
                num_centers, padded_centers = extract_centers(hm, thresh, radius=radius)
               
                for batch_idx, batch_centers in enumerate(padded_centers):
                    for center in batch_centers:
                        y = int(center[0])
                        x = int(center[1])
                        if x != 0 and y != 0:
                            maps[batch_idx, coord_range[2] + y, coord_range[0] + x] = 1

            # center_counts = torch.sum(maps, dim=(1, 2))
            # print('prev center count: ', center_counts)
            # Define the radius for combining centers

            # # Create a circular kernel for dilation
            # kernel_size = 2 * radius_duplicate + 1

            # # Use max pooling to find single centers in the combined area
            # pooled = F.max_pool2d(maps, kernel_size, stride=1, padding=radius_duplicate)

            # final_maps = maps * (pooled == maps)

            final_maps = remove_duplicate_centers(maps, radius_duplicate)

            num_centers = np.sum(final_maps, axis=(1, 2))

            #Print the sums for each batch
            # print("Sum of each batch:", num_centers)
            
            gt = gt_list[i]

            obj = objs[obj_idx]
            if obj in result_dict.keys():
                result_dict[obj]['gt_count'].extend(gt)
                result_dict[obj]['center_count'].extend(num_centers.tolist())
            else:
                result_dict[obj] = {'gt_count': gt, 'center_count': num_centers.tolist()}
        else:
            print('not exist')
        
        # if obj == 'pedestrian':
        #     print('gt count: {}, pred: {}'.format(gt, num_centers))

        if i % 2000 == 0:
            print(i)
    
    # eval2(result_dict)
    
    file = 'result_output/result_center_partition_multiscale_ep3_{}_{}_{}.txt'.format(thresh, radius, radius_duplicate)
    with open(file, 'w') as f: 
        json.dump(result_dict, f)
    return result_dict

def hm_partition_overlap(thresh, radius, radius_duplicate = 2):
    print('thresh: {}, radius: {}, radius duplicate: {}'.format(thresh, radius, radius_duplicate))
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    gt_path = 'gt_list.txt'
    if os.path.exists(gt_path):
            gt_list = json.load(open(gt_path))

    result_dict = {}
    expanding_rate = 0.2
    expanding = math.floor(64 * expanding_rate)
    # ranges = [[0, 64+expanding, 0, 64+expanding],[0, 64+expanding, 64-expanding, 128],[64-expanding, 128, 0, 64+expanding],[64-expanding, 128, 64-expanding, 128]]
    ranges = [[0, 64 + expanding, 0, 128], [64-expanding, 128, 0, 128]]

    # expanding = math.floor(42 * expanding_rate)
    # ranges = [[0, 42+expanding, 0, 42+expanding], [42-expanding, 84+expanding, 0, 42+expanding], [84-expanding, 128, 0, 42+expanding], [0, 42+expanding, 42-expanding, 84+expanding], [42-expanding, 84+ expanding, 42-expanding, 84+expanding], [84-expanding, 128, 42-expanding, 84+expanding], [0, 42+expanding, 84-expanding, 128], [42-expanding, 84+expanding, 84-expanding, 128], [84-expanding, 128, 84-expanding, 128]]
    
    start_time = datetime.now()
    for i in range(15050): #150
        obj_idx = i % 10
        file_path = '/data/hm_partition2_overlap0.3/{}.txt'.format(i)
        # file_path = '/data1/hms_partition_overlap_2/{}.txt'.format(i)
        # file_path = '/data1/hms_partition_overlap_0.3/{}.txt'.format(i)
        if os.path.exists(file_path):
            hms= json.load(open(file_path))
            # partition number of hms
            maps = np.zeros((4, 128, 128)) # batch size
            for hm_idx, hm in enumerate(hms):
                coord_range = ranges[hm_idx]
                hm= torch.tensor(hm, dtype=torch.float32)
                hm = sigmoid(hm)
                num_centers, padded_centers = extract_centers_dynamic2(hm, objs[obj_idx], thresh, radius=radius)
               
                for batch_idx, batch_centers in enumerate(padded_centers):
                    for center in batch_centers:
                        y = int(center[0])
                        x = int(center[1])
                        if x != 0 and y != 0:
                            maps[batch_idx, coord_range[2] + y, coord_range[0] + x] = 1

            final_maps = remove_duplicate_centers(maps, radius_duplicate)

            num_centers = np.sum(final_maps, axis=(1, 2))

            #Print the sums for each batch
            # print("Sum of each batch:", num_centers)
            
            gt = gt_list[i]

            obj = objs[obj_idx]
            if obj in result_dict.keys():
                result_dict[obj]['gt_count'].extend(gt)
                result_dict[obj]['center_count'].extend(num_centers.tolist())
            else:
                result_dict[obj] = {'gt_count': gt, 'center_count': num_centers.tolist()}
        else:
            print('not exist')
        
        # if obj == 'pedestrian':
        # print(obj)
        # print('gt count: {}, pred: {}'.format(gt, num_centers))

        if i % 1000 == 0 and i != 0:
            print(i)
            # end_time = datetime.now()
            # time_diff = (end_time - start_time).total_seconds()
            # print('time: {}'.format(time_diff/100))
            # break
    
    # eval2(result_dict)
    
    # file = 'result_output/result_center_partition_overlap_0.1_{}_{}_{}.txt'.format(thresh, radius, radius_duplicate)
    # # file = 'result_output/result_center_partition_overlap_ep3_dynamic_1.txt'.format(thresh)
    # # file = 'result_output/result_center_partition_overlap_2_dynamic.txt'.format(thresh)
    file = 'result_output/result_center_partition2_overlap_0.3_dynamic_{}.txt'.format(thresh)
    with open(file, 'w') as f: 
        json.dump(result_dict, f)
    return result_dict

def hm_partition(thresh, radius):
    print('thresh: {}, radius: {}'.format(thresh, radius))
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    gt_path = 'gt_list.txt'
    if os.path.exists(gt_path):
            gt_list = json.load(open(gt_path))

    result_dict = {}
    start_time = datetime.now()
    for i in range(15050): #15050
        obj_idx = i % 10
        # file_path = '/data/hms_partition_ep10/{}.txt'.format(i)
        # file_path = '/data/hms_partition_overlap9/{}.txt'.format(i)
        file_path = '/data1/hms_partition_2/{}.txt'.format(i)
        if os.path.exists(file_path):
            hms= json.load(open(file_path))
            # print(len(hms[0]))
            center_counts = torch.zeros(len(hms[0])) # batch size 4
            for hm in hms:
                hm= torch.tensor(hm, dtype=torch.float32)
                hm = sigmoid(hm)
                num_centers, padded_centers = extract_centers_dynamic(hm, thresh, radius=radius)

                if objs == 'motorcycle' or objs == 'bicycle':
                    num_centers, padded_centers = extract_centers_dynamic(hm, 0.12, radius=radius)
               
                center_counts += num_centers
                # print('hm shpe: ')
                # print(num_centers.shape)
                
            num_centers = center_counts.tolist()
            # print(num_centers)
            
            gt = gt_list[i]

            # if obj_idx == 0:
            #     print('gt {}, pred {}'.format(gt, num_centers))

            obj = objs[obj_idx]
            if obj in result_dict.keys():
                result_dict[obj]['gt_count'].extend(gt)
                result_dict[obj]['center_count'].extend(num_centers)
            else:
                result_dict[obj] = {'gt_count': gt, 'center_count': num_centers}
        else:
            print('not exist')
        
        # if obj == 'pedestrian':
        #     print('gt count: {}, pred: {}'.format(gt, num_centers))

        if i % 1000 == 0:
            print(i)
            # end_time = datetime.now()

            # time_diff = (end_time - start_time).total_seconds
            # print('time: {}'.format(time_diff/1000))
            # break
    
    # eval2(result_dict)
    
    file = 'result_output/result_center_partition_2_dynamic_{}_{}.txt'.format(thresh, radius)
    # file = 'result_output/result_center_partition_9_dynamic.txt'
    with open(file, 'w') as f: 
        json.dump(result_dict, f)
    return result_dict


def test_time_partition(thresh, radius, path):
    print('thresh: {}, radius: {}'.format(thresh, radius))
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    gt_path = 'gt_list.txt'
    if os.path.exists(gt_path):
            gt_list = json.load(open(gt_path))

    result_dict = {}
    start_time = datetime.now()
    for i in range(15050): #15050
        obj_idx = i % 10
        # file_path = '/data/hms_partition_ep10/{}.txt'.format(i)
        # file_path = '/data/hms_partition_overlap9/{}.txt'.format(i)
        file_path = path + '{}.txt'.format(i)
        if os.path.exists(file_path):
            hms= json.load(open(file_path))
            # print(len(hms[0]))
            center_counts = torch.zeros(len(hms[0])) # batch size 4
            for hm in hms:
                hm= torch.tensor(hm, dtype=torch.float32)
                hm = sigmoid(hm)
                num_centers, padded_centers = extract_centers_dynamic(hm, thresh, radius=radius)

                if objs == 'motorcycle' or objs == 'bicycle':
                    num_centers, padded_centers = extract_centers_dynamic(hm, 0.12, radius=radius)
               
                center_counts += num_centers
                # print('hm shpe: ')
                # print(num_centers.shape)
                
            num_centers = center_counts.tolist()
            # print(num_centers)
            
            gt = gt_list[i]

            # if obj_idx == 0:
            #     print('gt {}, pred {}'.format(gt, num_centers))

            obj = objs[obj_idx]
            if obj in result_dict.keys():
                result_dict[obj]['gt_count'].extend(gt)
                result_dict[obj]['center_count'].extend(num_centers)
            else:
                result_dict[obj] = {'gt_count': gt, 'center_count': num_centers}
        else:
            print('not exist')
        
        # if obj == 'pedestrian':
        #     print('gt count: {}, pred: {}'.format(gt, num_centers))

        if i % 1000 == 0 and i != 0:
            print(i)
            end_time = datetime.now()

            time_diff = (end_time - start_time).total_seconds()
            print('time: {}'.format(time_diff/1000))

            break

def eval2(result_dict):
    for key in result_dict.keys():
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']

        # print('pred:  {}, gt: {}'.format(pred, gt))
        # acc = accuracy_score(gt, pred)
        # print('category: {}, acc: {}'.format(key, acc))
        correct = 0
        print(key)
        
        diff = math.ceil(max(gt) * 0.1)
        # print('diff {}'.format(diff))
        for i in range(len(gt)):
            if gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff:
                correct += 1

        acc = correct/len(pred)
        print('category: {}, acc: {}'.format(key, acc))

def eval_combine(thresh, radius):
    object_combines = [['car', 'pedestrian'], ['car', 'barrier'], ['pedestrian', 'barrier'], ['car', 'pedestrian', 'barrier']]
    # objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    object_combines = [['truck', 'bus'], ['bus', 'trailer'], ['truck', 'trailer'], ['truck', 'bus', 'trailer']]

    #['truck', 'bus']
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_{}_{}_1.txt'.format(thresh, radius))
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_{}_{}.txt'.format(thresh, radius))

    # result_dict = load_eval_file('result_output/result_center_0.58_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_{}_{}_1.txt'.format(thresh, radius))
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/result_center_0.62_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    result_dict = load_eval_file('result_output/result_transfusion_0.02.txt')
    # result_dict = load_eval_file('result_output/result_voxelnext_0.2.txt')
    # result_dict = load_eval_file('result_output/result_center_pretrain_ep1_0.6_2.txt')
    result_dict = load_eval_file('result_output/result_safNet.txt')

    for combine in object_combines:
        frame_ids = len(result_dict['car']['center_count'])
        correct = 0

        for i in range(frame_ids):
            is_correct = True
            for item in combine:
                pred = result_dict[item]['center_count']
                gt = result_dict[item]['gt_count']

                diff = math.ceil(max(gt) * 0.1)
                
                if (gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff) == False:
                        is_correct = False
                
            if is_correct:
                correct += 1

        acc = correct/frame_ids
        print('category: {}, acc: {}'.format(combine, acc))

def eval_single_frame(model_name, frame_idx, objs):
    pred_list = []
    gt_list = []

    path = 'result_output/model_results/{}.txt'.format(model_name)
    result_dict = load_eval_file(path)
    # print(result_dict.keys())
    for key in objs:
        pred = result_dict[key]['center_count'][frame_idx]
        gt = result_dict[key]['gt_count'][frame_idx]

        pred_list.append(pred)
        gt_list.append(gt)
    return pred_list, gt_list

def getFeature():
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    model_list = ['partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2_overlap_0.3','partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4_overlap_0.3', 'partition_4', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9_overlap_0.3','partition_9'] #11
    model_list = ['partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9'] #12
    model_list = ['counternet', 'partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9'] #12
    model_list = ['partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9'] #15 add weight to obj
    # model_list = ['partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9']#14
    
    # model_list = ['partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4_overlap_0.3', 'partition_4', 'partition_9_overlap_0.2', 'partition_9'] #10
    # model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_4_overlap_0.3', 'partition_9'] #7
    # model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_4', 'partition_4_overlap_0.3', 'partition_9'] #6
    # model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_9_overlap_0.2', 'partition_9'] #8
    # model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_9_overlap_0.2'] #9

    model_frame_dict = {string: [] for string in model_list}

    frame_idx = 0
    for i in range(14000): #15050
        # for each i, find 10 obj in each model
        file_path = '/data/feature_nus/' + '{}.txt'.format(i)
        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                batch_tensor = pickle.load(f)
                # print(batch_tensor.shape)
                
                for tensor in batch_tensor:
                    
                    selected_model_idx = 0
                    min_diff = 1000
                    for m_idx, model in enumerate(model_list):
                        # get model
                        try:
                            preds, gts = eval_single_frame(model, frame_idx, objs) # expect to return count of all object and the corresponding gt in list
                        except IndexError:
                            break
                        # print(preds, gts)
                        # diff = sum(abs(a - b) for a, b in zip(preds, gts))

                        diff = 0
                        weight = [0.2, 0.1 ,0.1 ,0.1 ,0.1 , 0.1, 0.1 , 0.1, 0.2, 0.1]
                        for idx, obj in enumerate(objs):
                            diff += weight[idx] * abs(preds[idx] - gts[idx])

                        
                        # print('frame_id: {}, model: {}, diff: {}'.format(frame_idx, model, diff))
                        if diff < min_diff:
                            selected_model_idx = m_idx
                            min_diff = diff
                    model_frame_dict[model_list[selected_model_idx]].append(tensor.tolist())    
                    frame_idx += 1
                    
    for key in model_frame_dict.keys():
        print('key: {}, length: {}'.format(key, len(model_frame_dict[key])))
    # print(model_frame_dict)
    file = 'result_output/model_results/model_frame_dict15.txt'
    with open(file, 'w') as f: 
        json.dump(model_frame_dict, f)

def getCenter():
    center_dict = {}
    file_path = 'result_output/model_results/model_frame_dict12.txt'
    if os.path.exists(file_path):
            model_frame_dict = json.load(open(file_path))
            for key in model_frame_dict.keys():
                data_array = np.array(model_frame_dict[key])
                avg_pool = np.mean(data_array, axis=0)
                
                distance_sum = 0

                for array in data_array:
                    distance = np.linalg.norm(array - avg_pool)
                    distance_sum += distance
                avg_distance = distance_sum/len(data_array) # sigma square
                print(avg_distance)

                # adjustment chernoff bound
                epsilon = 0.2
                # print('key: {}, avg_dist: {}'.format(key, avg_distance))
                power = -(len(data_array) * epsilon*epsilon) / (2 * avg_distance)
                adjustment = math.exp(power)
                print('key: {}, adjustment: {}'.format(key, adjustment))
                # model_frame_dict[key] = avg_pool.tolist()

                if key not in center_dict.keys():
                    center_dict[key] = {'center': avg_pool, 'adjustment': adjustment}
    return center_dict

def epsilonSta():
    center_dict = {}
    file_path = 'result_output/model_results/model_frame_dict12.txt'
    eps = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
    if os.path.exists(file_path):
            model_frame_dict = json.load(open(file_path))
            output_dict = {}
            for ep in eps:
                for key in model_frame_dict.keys():
                    data_array = np.array(model_frame_dict[key])
                    avg_pool = np.mean(data_array, axis=0)
                    
                    distance_sum = 0

                    for array in data_array:
                        distance = np.linalg.norm(array - avg_pool)
                        distance_sum += distance
                    avg_distance = distance_sum/len(data_array) # sigma square
                    print(avg_distance)

                    # adjustment chernoff bound
                    epsilon = ep
                    # print('key: {}, avg_dist: {}'.format(key, avg_distance))
                    power = -(len(data_array) * epsilon*epsilon) / (2 * avg_distance)
                    adjustment = math.exp(power)
                    print('key: {}, adjustment: {}'.format(key, adjustment))
                    
                    if key in output_dict.keys():
                        output_dict[key].append(adjustment)
                    else:
                        output_dict[key] = [adjustment]
                    # model_frame_dict[key] = avg_pool.tolist()

                    if key not in center_dict.keys():
                        center_dict[key] = {'center': avg_pool, 'adjustment': adjustment}

            print(output_dict)
    return center_dict

def evalModelSelection():
    center_dict = getCenter()
    output_dict = {}
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    # model_list = ['partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4_overlap_0.3', 'partition_4', 'partition_9_overlap_0.2', 'partition_9'] #10
    # model_list = ['partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2_overlap_0.3','partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4_overlap_0.3', 'partition_4', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9_overlap_0.3','partition_9'] #11
    # model_list = ['partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9'] #12
    model_list = ['counternet', 'partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9'] #13
    model_list = ['partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9'] #14
    model_list = ['partition_2_overlap_0.1', 'partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4', 'partition_9_overlap_0.1','partition_9_overlap_0.2', 'partition_9']
    # model_list = ['partition_4_overlap_0.2', 'partition_4', 'partition_9_overlap_0.2', 'partition_9']
    # model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_4', 'partition_4_overlap_0.3', 'partition_9'] #6
    # model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_4_overlap_0.3', 'partition_9'] # 7
    # model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_9_overlap_0.2', 'partition_9'] #8
    # model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_9_overlap_0.2'] #9

    frame_idx = 0
    model_count = {}
    for i in range(15050): #15050
        # for each i, find 10 obj in each model
        file_path = '/data/feature_nus/' + '{}.txt'.format(i)
        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                batch_tensor = pickle.load(f)
                # print(batch_tensor.shape)
                for tensor in batch_tensor:
                    selected_model = model_list[0]
                    distance_min = 1000
                    for m_idx, model in enumerate(model_list):
                        center = center_dict[model]['center'] #'adjustment'
                        adjustment = center_dict[model]['adjustment']
                        distance = np.linalg.norm(center - tensor.tolist())
                        distance += adjustment * distance
                        # print('model: {}, adjust: {}, distance: {}'.format(model, adjustment, distance))
                        if distance < distance_min:
                            distance_min = distance
                            selected_model = model

                    if selected_model not in model_count.keys():
                        model_count[selected_model] = 1
                    else:
                        model_count[selected_model] += 1


                    try:
                        preds, gts = eval_single_frame(selected_model, frame_idx, objs) # expect to return count of all object and the corresponding gt in list
                    except IndexError:
                        break

                    for idx, pred in enumerate(preds):
                        obj = objs[idx]
                        if obj not in output_dict.keys():
                            output_dict[obj] = {'gt_count': [gts[idx]], 'center_count':[pred]}
                        else:
                            output_dict[obj]['gt_count'].append(gts[idx])
                            output_dict[obj]['center_count'].append(pred)
                    frame_idx += 1
    # print(frame_idx)
    # print(len(output_dict['pedestrian']['gt_count']))
    print(model_count)
    file = 'result_output/model_results/result_model_selection15_car_ped_adjustment_0.3.txt'
    with open(file, 'w') as f: 
        json.dump(output_dict, f)

def evalRecall():
    # result_dict = load_eval_file('result_output/result_center_0.6_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_{}_{}_1.txt'.format(thresh, radius))
    result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_voxelnext_0.3.txt')
    # result_dict = load_eval_file('result_output/result_transfusion_0.5.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap9_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.12.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_0.3_0.12_2_1.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_2_dynamic_0.12.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_0.1_0.07_2_1.txt')
    # model_list = ['partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4_overlap_0.3', 'partition_4', 'partition_9_overlap_0.2', 'partition_9']
    # result_dict = load_eval_file('result_output/model_results/{}.txt'.format(model_list[4]))
    # result_dict = load_eval_file('result_output/result_center_partition9_overlap_0.3_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/model_results/result_model_selection13.txt')
    # result_dict = load_eval_file('result_output/kitti/result_counternet2_0.42_10.txt')
    # result_dict = load_eval_file('result_output/kitti_sta_center_point_0.3.txt')
    result_dict = load_eval_file('result_output/result_safNet.txt')

    obj_list = []
    acc_list = []

    print(result_dict.keys())
    for key in result_dict.keys():
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']

        # print('pred:  {}, gt: {}'.format(pred, gt))
        # acc = accuracy_score(gt, pred)
        # print('category: {}, acc: {}'.format(key, acc))
        correct = 0
        all = 0
        print(key)
        
        diff = math.ceil(max(gt) * 0.05)
        print('diff {}'.format(diff))
        for i in range(len(gt)):
            if gt[i] != 0:
                all += 1
                if gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff:
                    correct += 1

        acc = correct/all #len(pred)
        print('category: {}, recall: {}'.format(key, acc))
        
        obj_list.append(key)
        acc_list.append(acc)

    return obj_list,acc_list


def eval(thresh, radius, ep=0.1):
    result_dict = load_eval_file('result_output/result_center_0.62_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_{}_{}_1.txt'.format(thresh, radius))
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.12.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_voxelnext_0.2.txt')
    # result_dict = load_eval_file('result_output/result_transfusion_0.02.txt')
    # # result_dict = load_eval_file('result_output/result_transfusion_0.02.txt')
    result_dict = load_eval_file('result_output/result_center_partition_overlap9_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.12.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_0.3_0.12_2_1.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_2_dynamic_0.12.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_0.1_0.07_2_1.txt')
    # model_list = ['partition_2_overlap_0.2', 'partition_2', 'partition_4_overlap_0.1', 'partition_4_overlap_0.2', 'partition_4_overlap_0.3', 'partition_4', 'partition_9_overlap_0.2', 'partition_9']
    # result_dict = load_eval_file('result_output/model_results/{}.txt'.format(model_list[4]))
    # result_dict = load_eval_file('result_output/result_center_partition9_overlap_0.3_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/model_results/result_model_selection14_adjustment.txt')
    # result_dict = load_eval_file('result_output/kitti/result_counternet2_23.9_10.txt')
    # result_dict = load_eval_file('result_output/kitti/result_counternet2_23.9_10.txt')
    # result_dict = load_eval_file('result_output/kitti_sta_center_point_0.28.txt')


    #13: 9 models
    #14: 6 models
    #15: car_ped
    # if ep == 0:
    #     result_dict = load_eval_file('result_output/model_results/result_model_selection15_car_ped.txt')
    # else:
    #     result_dict = load_eval_file('result_output/model_results/result_model_selection15_car_ped_adjustment_{}.txt'.format(ep))
    if ep == 0:
        result_dict = load_eval_file('result_output/model_results/result_model_selection14.txt')
    else:
        result_dict = load_eval_file('result_output/model_results/result_model_selection14_adjustment_{}.txt'.format(ep))

    # result_dict = load_eval_file('result_output/result_center_partition_overlap_2_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/model_results/partition_4_overlap_0.1.txt')
    # result_dict = load_eval_file('result_output/kitti_sta_center_point_0.3.txt')
    
    
    result_dict = load_eval_file('result_output/result_safNet.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_0.12_2_1.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_9_dynamic_0.12_2.txt')
    # result_dict = load_eval_file('result_output/result_center_pretrain_ep1_0.6_2.txt')
    # print(result_dict.keys())
    obj_list = []
    acc_list = []

    print(result_dict.keys())
    for key in result_dict.keys():
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']

        # print('pred:  {}, gt: {}'.format(pred, gt))
        # acc = accuracy_score(gt, pred)
        # print('category: {}, acc: {}'.format(key, acc))
        correct = 0
        # print(key)
        
        diff = math.ceil(max(gt) * 0.05)
        print('diff {}'.format(diff))
        for i in range(len(gt)):
            if gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff:
                correct += 1

        acc = correct/len(pred)
        print('category: {}, acc: {}'.format(key, acc))
        
        obj_list.append(key)
        acc_list.append(acc)

    return obj_list,acc_list

def eval_binary(thresh, radius):
    # result_dict = load_eval_file('result_output/result_center_0.62_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_{}_{}_1.txt'.format(thresh, radius))
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_voxelnext_0.2.txt')
    result_dict = load_eval_file('result_output/result_transfusion_0.02.txt')
    # result_dict = load_eval_file('result_output/result_center_pretrain_ep1_0.6_2.txt')
    # print(result_dict.keys())
    result_dict = load_eval_file('result_output/result_safNet.txt')
    obj_list = []
    acc_list = []

    print(result_dict.keys())
    for key in result_dict.keys():
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']

        # print('pred:  {}, gt: {}'.format(pred, gt))
        # acc = accuracy_score(gt, pred)
        # print('category: {}, acc: {}'.format(key, acc))
        correct = 0
        total = 0
        print(key)
        
        diff = 0 #math.ceil(max(gt) * 0.05)
        # print('diff {}'.format(diff))
        # for i in range(len(gt)):
        #     if gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff:
        #         correct += 1

        for i in range(len(gt)):
            if gt[i] != 0 and pred[i] != 0:
                correct += 1
            if gt[i] == 0 and pred[i] == 0:
                correct += 1

        print('total {}'.format(total))
        acc = correct/len(pred)
        print('category: {}, acc: {}'.format(key, acc))
        
        obj_list.append(key)
        acc_list.append(acc)

    return obj_list,acc_list

def eval_count(thresh, radius):
    # result_dict = load_eval_file('result_output/result_center_0.62_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_{}_{}_1.txt'.format(thresh, radius))
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_voxelnext_0.2.txt')
    result_dict = load_eval_file('result_output/result_transfusion_0.02.txt')
    # result_dict = load_eval_file('result_output/result_center_pretrain_ep1_0.6_2.txt')
    # print(result_dict.keys())
    result_dict = load_eval_file('result_output/result_safNet.txt')
    obj_list = []
    acc_list = []

    print(result_dict.keys())
    for key in result_dict.keys():
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']

        # print('pred:  {}, gt: {}'.format(pred, gt))
        # acc = accuracy_score(gt, pred)
        # print('category: {}, acc: {}'.format(key, acc))
        correct = 0
        total = 0
        print(key)
        
        diff = math.ceil(max(gt) * 0.05)
        # print('diff {}'.format(diff))
        # for i in range(len(gt)):
        #     if gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff:
        #         correct += 1
        
        total_acc = 0
        for j in range(1000):
            correct = 0
            total = 0
            random_number = random.randint(0, len(gt) - 1 - 250)
            for i in range(random_number, random_number + 250 +1):
                if gt[i] == 5:
                    total += 1
                    # if gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff:
                    if gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff:
                        correct += 1
            if total == correct:
                total_acc += 1
            else:
                total_acc += correct/total


        acc = total_acc/1000 #correct/len(pred)
        print('category: {}, acc: {}'.format(key, acc))
        
        obj_list.append(key)
        acc_list.append(acc)

    return obj_list,acc_list

def eval_agg(thresh, radius):
    # result_dict = load_eval_file('result_output/result_center_0.62_2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_{}_{}_1.txt'.format(thresh, radius))
    result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_voxelnext_0.2.txt')
    # result_dict = load_eval_file('result_output/result_transfusion_0.02.txt')
    # result_dict = load_eval_file('result_output/result_center_pretrain_ep1_0.6_2.txt')
    # print(result_dict.keys())
    result_dict = load_eval_file('result_output/result_safNet.txt')
    obj_list = []
    acc_list = []

    print(result_dict.keys())
    for key in result_dict.keys():
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']

        # print('pred:  {}, gt: {}'.format(pred, gt))
        # acc = accuracy_score(gt, pred)
        # print('category: {}, acc: {}'.format(key, acc))
        correct = 0
        total = 0
        print(key)
        
        diff = math.ceil(max(gt) * 0.05)
        # print('diff {}'.format(diff))
        # for i in range(len(gt)):
        #     if gt[i] >= pred[i] - diff and gt[i] <= pred[i] + diff:
        #         correct += 1
        
        total_diff_avg = 0
        for j in range(1000):
            gt_sum = 0
            pred_sum = 0
            random_number = random.randint(0, len(gt) - 1 - 250)
            for i in range(random_number, random_number + 250 +1):
                gt_sum += gt[i]
                pred_sum += pred[i]


            # if total == correct:
            #     total_acc += 1
            # else:
            #     total_acc += correct/total
            total_diff_avg += abs(gt_sum - pred_sum)/250


        acc = total_diff_avg/1000
        print('category: {}, acc: {}'.format(key, acc))
        
        obj_list.append(key)
        acc_list.append(acc)

    return obj_list,acc_list

def eval_agg_q_error(thresh, radius):
    # result_dict = load_eval_file('result_output/result_center_0.62_2.txt')
    # result_dict = load_eval_file('result_output/model_results/partition_4.txt')
    result_dict = load_eval_file('result_output/model_results/partition_4_overlap_0.2.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_{}_{}_1.txt'.format(thresh, radius))
    # result_dict = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.07.txt')
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_dynamic_0.07_2.txt')
    # result_dict = load_eval_file('result_output/result_voxelnext_0.2.txt')
    # result_dict = load_eval_file('result_output/result_transfusion_0.02.txt')
    # result_dict = load_eval_file('result_output/result_center_pretrain_ep1_0.6_2.txt')
    # print(result_dict.keys())
    result_dict = load_eval_file('result_output/result_safNet.txt')
    obj_list = []
    acc_list = []

    print(result_dict.keys())
    for key in result_dict.keys():
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']

        print(key)

        total_diff_avg = []

        for j in range(1000):
            q_errors = []
            gt_sum = []
            pred_sum = []
            random_number = random.randint(0, len(gt) - 1 - 250)
            for i in range(random_number, random_number + 250 +1):
                gt_sum.append(gt[i])
                pred_sum.append(pred[i])

            
            # for t, e in zip(gt_sum, pred_sum):
            #     if t == 0 and e == 0:
            #         q_errors.append(1.0)
            #     elif t == 0 or e == 0:
            #         t += 1
            #         e += 1  
            #         q_errors.append(np.maximum(t / e, e / t))
            #     else:
            #         q_errors.append(np.maximum(t / e, e / t))

            gt1 = sum(gt_sum) + 1
            pred1 = sum(pred_sum) + 1
            q_errors = np.maximum(gt1/ pred1, pred1 / gt1)
            total_diff_avg.append(q_errors) #np.median(q_errors)
            # print("Median Q-error:", np.median(q_errors))
            # print('mean q-error:', np.mean(q_errors))
        acc = np.median(q_errors)
        print('category: {}, acc: {}'.format(key, acc))
        
        obj_list.append(key)
        acc_list.append(acc)

    return obj_list,acc_list


def getData():
    # threshs = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6]
    threshs = [0.5, 0.52, 0.54, 0.55, 0.56, 0.58, 0.6, 0.62]
    radius = 2
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    output_dict = {} #{'car': [result_0.1, result_0.2, ...]}
    for thresh in threshs:
        obj_list, acc_list = eval(thresh, 2)
        for i in range(len(obj_list)):
            obj = obj_list[i]
            acc = acc_list[i]

            if obj in output_dict.keys():
                output_dict[obj].append(acc)
            else:
                output_dict[obj] = [acc]
    print(output_dict)

def getData2():
    # threshs = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6]
    eps = [0, 0.1, 0.15, 0.2, 0.3]
    radius = 2
    objs = ['car', 'truck', 'bus','trailer','construction_vehicle', 'pedestrian',  'motorcycle', 'bicycle',  'traffic_cone', 'barrier']
    output_dict = {} #{'car': [result_0.1, result_0.2, ...]}
    for thresh in eps:
        obj_list, acc_list = eval(0.07, 2, thresh)
        for i in range(len(obj_list)):
            obj = obj_list[i]
            acc = acc_list[i]

            if obj in output_dict.keys():
                output_dict[obj].append(acc)
            else:
                output_dict[obj] = [acc]
    print(output_dict)

    return output_dict
def write_to_file(result_dict, sheet_name):
    objs = ['car', 'truck', 'bus','trailer','construction_vehicle', 'pedestrian',  'motorcycle', 'bicycle',  'traffic_cone', 'barrier']
    file_name = 'adjustment_6.xlsx'
    if not os.path.isfile(file_name):
        wb = Workbook()
    else:
        wb = load_workbook(file_name)

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)
        sheet.title = sheet_name

    sheet.cell(row=1, column=column_index_from_string('A')).value = 'Category'
    sheet.cell(row=1, column=column_index_from_string('B')).value = '0'
    sheet.cell(row=1, column=column_index_from_string('C')).value = '0.1'
    sheet.cell(row=1, column=column_index_from_string('D')).value = '0.15'
    sheet.cell(row=1, column=column_index_from_string('E')).value = '0.2'
    sheet.cell(row=1, column=column_index_from_string('F')).value = '0.3'
    # sheet.cell(row=1, column=column_index_from_string('G')).value = '0.58'
    # sheet.cell(row=1, column=column_index_from_string('H')).value = '0.6'
    # sheet.cell(row=1, column=column_index_from_string('I')).value = '0.62'
    # sheet.cell(row=1, column=column_index_from_string('J')).value = '0.59'
    # sheet.cell(row=1, column=column_index_from_string('K')).value = '0.6'


    row = sheet.max_row +1
    # print('row_number: ', row)
    for i in range(len(objs)):
        acc_list = result_dict[objs[i]]
        sheet.cell(row=row, column=column_index_from_string('A')).value = objs[i]
        sheet.cell(row=row, column=column_index_from_string('B')).value = acc_list[0]
        sheet.cell(row=row, column=column_index_from_string('C')).value = acc_list[1]
        sheet.cell(row=row, column=column_index_from_string('D')).value = acc_list[2]
        sheet.cell(row=row, column=column_index_from_string('E')).value = acc_list[3]
        sheet.cell(row=row, column=column_index_from_string('F')).value = acc_list[4]
        # sheet.cell(row=row, column=column_index_from_string('F')).value = acc_list[4]
        # sheet.cell(row=row, column=column_index_from_string('G')).value = acc_list[5]
        # sheet.cell(row=row, column=column_index_from_string('H')).value = acc_list[6]
        # sheet.cell(row=row, column=column_index_from_string('I')).value = acc_list[7]
        # sheet.cell(row=row, column=column_index_from_string('J')).value = acc_list[8]
        # sheet.cell(row=row, column=column_index_from_string('K')).value = acc_list[9]

        row += 1

    # sheet.cell(row=row, column=column_index_from_string('O')).value = time

    wb.save(file_name)

def map_values(values, ranges):
    mapped_values = []
    for value in values:
        for key, (low, high) in ranges.items():
            if low <= value <= high:
                mapped_values.append(key)
                break
    return mapped_values
def processWithRange(ground_truth, predictions):
    ranges = {
        0:(0, 5),
        6:(6, 10),
        11:(11, 15),
        15:(15, 20),
        16:(16, 20),
        21:(21, 30),
        31:(31, 40),
        41:(41, 50),
        51:(51, 100),
    }
    mapped_ground_truth = map_values(ground_truth, ranges)
    mapped_predictions = map_values(predictions, ranges)

    return mapped_ground_truth, mapped_predictions

def getCategoricalAcc(ground_truth, predictions):
    # Get unique categories
    categories = np.unique(ground_truth + predictions)

    # Calculate confusion matrix
    conf_matrix = confusion_matrix(ground_truth, predictions, labels=categories)

    # Calculate accuracy for each category
    category_accuracies = {}
    for i, category in enumerate(categories):
        true_positives = conf_matrix[i, i]
        total_samples = sum(conf_matrix[i, :])
        accuracy = true_positives / total_samples if total_samples != 0 else 0
        category_accuracies[category] = accuracy

    return category_accuracies
    # Display the accuracy for each category
    # for category, accuracy in category_accuracies.items():
    #     print(f"Count {category}: Accuracy {accuracy:.2f}")

def testRange():
    # result_dict = load_eval_file('result_output/result_center_partition_ep10_0.57_2.txt')
    result_dict = load_eval_file('result_output/result_center_0.6_2.txt')
    for key in result_dict.keys():
        
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']
        

        print(key)
        gt, pred = processWithRange(gt, pred)
        print(getCategoricalAcc(gt, pred))

def test():
    model_list = ['partition_2', 'partition_4_overlap_0.2', 'partition_4', 'partition_9_overlap_0.2']
    result_dict = load_eval_file('result_output/model_results/{}.txt'.format(model_list[3]))

    print('gt: ')
    print(result_dict['car']['gt_count'][:10])
    print('pred: ')
    print(result_dict['car']['center_count'][:10])

def combine():
    output_dict = {}
    
    objs = ['car', 'truck', 'construction_vehicle', 'bus', 'trailer', 'barrier', 'motorcycle', 'bicycle', 'pedestrian', 'traffic_cone']
    # result_dict1 = load_eval_file('result_output/result_center_partition2_overlap_0.3_dynamic_0.07.txt')   
    # result_dict2 = load_eval_file('result_output/result_center_partition2_overlap_0.3_dynamic_0.12.txt') 
    # result_dict1 = load_eval_file('result_output/result_center_partition_2_dynamic_0.07_2.txt')
    # result_dict2 = load_eval_file('result_output/result_center_partition_2_dynamic_0.12_2.txt')
    # result_dict1 = load_eval_file('result_output/result_center_partition_overlap_2_dynamic_0.07.txt')
    # result_dict2 = load_eval_file('result_output/result_center_partition_overlap_2_dynamic_0.12.txt')
    # result_dict1 = load_eval_file('result_output/result_center_partition2_overlap_0.1_dynamic_0.07.txt')
    # result_dict2 = load_eval_file('result_output/result_center_partition2_overlap_0.1_dynamic_0.12.txt')
    # result_dict1 = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.07.txt')
    # result_dict2 = load_eval_file('result_output/result_center_partition_overlap_ep3_dynamic_0.12.txt')
    # result_dict1 = load_eval_file('result_output/result_center_partition_overlap_0.3_0.07_2_1.txt')
    # result_dict2 = load_eval_file('result_output/result_center_partition_overlap_0.3_0.12_2_1.txt')
    result_dict1 = load_eval_file('result_output/result_safNet_0.02.txt')
    result_dict2 = load_eval_file('result_output/result_safNet_0.03.txt')
    for obj in objs:
        # if obj == 'construction_vehicle' or obj == 'motorcycle' or obj == 'bicycle':
        if obj == 'car' or obj == 'truck' or obj == 'bus':
            output_dict[obj] = result_dict2[obj]
        else:
            output_dict[obj] = result_dict1[obj]
    file = 'result_output/result_safNet.txt'
    with open(file, 'w') as f: 
        json.dump(output_dict, f)
    

def getSelectivity():
    result_dict = load_eval_file('result_output/model_results/partition_9_overlap_0.1.txt')
    for key in result_dict.keys():
        pred = result_dict[key]['center_count']
        gt = result_dict[key]['gt_count']

        selected_count = 0
        if key == 'barrier':
            for item in gt:
                if item <3 and item >0:
                    selected_count += 1
        print('selectivity of {} is {}'.format(key, selected_count/len(gt)))

def getSelectivityCombine():
    combine = ['truck', 'barrier']
    result_dict = load_eval_file('result_output/model_results/counternet.txt')

    frame_ids = len(result_dict['car']['center_count'])
    correct = 0

    gt1 = result_dict[combine[0]]['gt_count']
    gt2 = result_dict[combine[1]]['gt_count']
    
    for i in range(frame_ids):
        # print('len of gt1 {}, len of gt2 {}'.format(len(gt1), len(gt2)))
        # print(i)
        if gt1[i] > 0 and gt2[i] > 5:
            # print(i)
            correct += 1

    selectivity = correct/frame_ids
    print('category: {}, selectivity: {}'.format(combine, selectivity))

def evalSelectivity():
    combine = ['truck', 'barrier']
    result_dict = load_eval_file('result_output/model_results/counternet.txt')
    result_dict = load_eval_file('result_output/model_results/partition_4_overlap_0.2.txt')
    # result_dict = load_eval_file('result_output/model_results/partition_4.txt')
    result_dict = load_eval_file('result_output/result_voxelnext_0.2.txt')
    result_dict = load_eval_file('result_output/result_transfusion_0.02.txt')

    frame_ids = len(result_dict['car']['gt_count'])

    gt_list = []
    pred_list = []

    gt1 = result_dict[combine[0]]['gt_count']
    gt2 = result_dict[combine[1]]['gt_count']

    pred1 = result_dict[combine[0]]['center_count']
    pred2 = result_dict[combine[1]]['center_count']

    for i in range(frame_ids):
        # print('len of gt1 {}, len of gt2 {}'.format(len(gt1), len(gt2)))
        # print(i)
        if gt1[i] > 0 and gt2[i] > 5:
            gt_list.append(i)
        if pred1[i] > 0 and pred2[i] > 5:
            pred_list.append(i)

    same_items = len(set(gt_list) & set(pred_list))
    print('gt list {}, pred list {}'.format(len(gt_list), len(pred_list)))
    # print('acc: {}'.format(same_items/len(set(gt_list).union(set(pred_list)))))
    recall = same_items/len(pred_list)
    precision = same_items/len(gt_list)
    acc = same_items/len(set(gt_list).union(set(pred_list)))
    print('precision: {}, recall: {}, acc: {}'.format(precision, recall, acc))

    # selectivity = correct/frame_ids
    # print('category: {}, selectivity: {}'.format(combine, selectivity))
    
# testRange()
# evalRecall()
# eval(0.58, 2)
# getSelectivity()

# evalSelectivity()

# getSelectivityCombine()
# combine()
# test()
# eval_binary(0.58, 2)
# eval_count(0.58, 2)
# eval_agg(0.58, 2)
# eval_agg_q_error(0.58, 2)
eval_combine(0.58, 2)
# result_dict = getData2()
# write_to_file(result_dict, 'thresh_0.2')
# # hm()

# getFeature()
# getCenter()
# epsilonSta()
# evalModelSelection()

# test_time_partition(0.05, 2, '/data1/hms_partition_2/')
# test_time_partition(0.05, 2, '/data/hms_partition_ep10/')
# test_time_partition(0.05, 2, '/data/hms/')
# hm_partition(0.05, 2)
# hm_partition(0.06, 2)
# hm_partition(0.07, 2)
# hm_partition(0.08, 2)
# hm_partition(0.09, 2)
# hm_partition(0.1, 2)
# hm_partition(0.11, 2)
# hm_partition(0.12, 2)

# hm_partition_overlap(0.07, 2, 1)
# hm_partition_overlap(0.12, 2, 1)
# hm_partition_overlap(0.09, 2, 1)
# hm_partition_overlap(0.08, 2, 1)
# hm_partition_overlap(0.11, 2, 1)
# hm_partition_overlap(0.1, 2, 1)
# hm_partition_overlap(0.06, 2, 1)
# hm_partition_overlap(0.05, 2, 1)

# hm_partition(0.6, 2)
# hm_partition(0.59, 2)
# hm_partition(0.58, 2)
# hm_partition(0.56, 2)
# hm_partition(0.55, 2)
# hm_partition(0.54, 2)
# hm_partition(0.53, 2)
# hm_partition(0.52, 2)
# hm_partition(0.51, 2)
# hm_partition(0.50, 2)
# hm_partition(0.62, 2)
# hm_partition(0.61, 2)
# hm_partition(0.62, 3)
# hm_partition(0.61, 3)

# hm_partition_overlap(0.61, 2, 3)
# hm_partition_overlap(0.59, 2, 3)
# hm_partition_overlap(0.63, 2, 3)

# hm_partition_overlap(0.09, 2, 1)
# hm_partition_overlap(0.08, 2, 1)
# hm_partition_overlap(0.07, 2, 1)
# hm_partition_overlap(0.11, 2, 1)
# hm_partition_overlap(0.12, 2, 1)
# hm_partition_overlap(0.06, 2, 1)

# hm_partition_overlap(0.57, 2, 1)
# hm_partition_overlap(0.57, 2, 2)
# hm_partition_overlap(0.57, 2, 3)

# hm_partition_overlap(0.575, 2, 1)
# hm_partition_overlap(0.575, 2, 2)
# hm_partition_overlap(0.575, 2, 3)

# hm_partition_overlap(0.585, 2, 1)
# hm_partition_overlap(0.585, 2, 2)
# hm_partition_overlap(0.585, 2, 3)

# hm_partition_overlap(0.61, 2, 2)
# hm_partition_overlap(0.59, 2, 2)
# hm_partition_overlap(0.63, 2, 2)

# hm_partition_overlap(0.6, 2, 1)
# hm_partition_overlap(0.62, 2, 1)
# hm_partition_overlap(0.58, 2, 1)
# hm_partition_overlap(0.54, 2, 1)
# hm_partition_overlap(0.52, 2, 1)
# hm_partition_overlap(0.55, 2, 1)
# hm_partition_overlap(0.56, 2, 1) 
# hm_partition_overlap(0.5, 2, 1)

# hm_partition_overlap(0.6, 2, 2)
# hm_partition_overlap(0.62, 2, 2)
# hm_partition_overlap(0.58, 2, 2)
# hm_partition_overlap(0.56, 2, 2)
# hm_partition_overlap(0.55, 2, 2)
# hm_partition_overlap(0.54, 2, 2)
# hm_partition_overlap(0.52, 2, 2)
# hm_partition_overlap(0.5, 2, 2)

# hm_partition_overlap(0.6, 2, 0)
# hm_partition_overlap(0.62, 2, 0)
# hm_partition_overlap(0.58, 2, 0)
# hm_partition_overlap(0.56, 2, 0)
# hm_partition_overlap(0.55, 2, 0)
# hm_partition_overlap(0.54, 2, 0)
# hm_partition_overlap(0.52, 2, 0)
# hm_partition_overlap(0.5, 2, 0)

# hm_partition_overlap(0.6, 2, 3)
# hm_partition_overlap(0.62, 2, 3)
# hm_partition_overlap(0.58, 2, 3)
# hm_partition_overlap(0.56, 2, 3)
# hm_partition_overlap(0.55, 2, 3)
# hm_partition_overlap(0.54, 2, 3)
# hm_partition_overlap(0.52, 2, 3)
# hm_partition_overlap(0.5, 2, 3)

# hm(0.05, 2)
# hm(0.1, 2)
# hm(0.49, 10)
# hm(24, 10)
# hm(24.2, 10)
# hm(24.4, 10)
# hm(24.6, 10)
# hm(23, 10)
# hm(23.5, 10)
# hm(23.9, 10)
# hm(23.8, 10)
# hm(23.7, 10)
# hm(0.4, 10)
# # hm(0.48, 5)
# # hm(0.46, 5)
# hm(0.42, 10)
# hm(0.44, 10)

# hm(0.4, 5)
# hm(0.34, 5)
# hm(0.33, 5)
# hm(0.32, 5)
# hm(0.31, 5)
# hm(0.3, 5)




# hm(0.58, 2)
# hm(0.59, 2)
# hm(0.61, 2)
# hm(0.62, 2)
# hm(0.63, 2)
# hm(0.64, 2)
# hm(0.65, 2)
# hm(0.66, 2)

# hm(0.61, 3)
# hm(0.62, 3)





