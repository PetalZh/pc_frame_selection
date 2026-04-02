import pickle
import time

import numpy as np
import torch
import tqdm

from pcdet.models import load_data_to_gpu
from pcdet.utils import common_utils
from pcdet.constants import result as count_result
from pcdet.constants import gt_list, sample_method, sample_rate
import json, os
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import Workbook



def statistics_info(cfg, ret_dict, metric, disp_dict):
    for cur_thresh in cfg.MODEL.POST_PROCESSING.RECALL_THRESH_LIST:
        metric['recall_roi_%s' % str(cur_thresh)] += ret_dict.get('roi_%s' % str(cur_thresh), 0)
        metric['recall_rcnn_%s' % str(cur_thresh)] += ret_dict.get('rcnn_%s' % str(cur_thresh), 0)
    metric['gt_num'] += ret_dict.get('gt', 0)
    min_thresh = cfg.MODEL.POST_PROCESSING.RECALL_THRESH_LIST[0]
    disp_dict['recall_%s' % str(min_thresh)] = \
        '(%d, %d) / %d' % (metric['recall_roi_%s' % str(min_thresh)], metric['recall_rcnn_%s' % str(min_thresh)], metric['gt_num'])


def eval_one_epoch(cfg, args, model, dataloader, epoch_id, logger, dist_test=False, result_dir=None):
    # start_time = time.time()
    result_dir.mkdir(parents=True, exist_ok=True)

    final_output_dir = result_dir / 'final_result' / 'data'
    if args.save_to_file:
        final_output_dir.mkdir(parents=True, exist_ok=True)

    metric = {
        'gt_num': 0,
    }
    for cur_thresh in cfg.MODEL.POST_PROCESSING.RECALL_THRESH_LIST:
        metric['recall_roi_%s' % str(cur_thresh)] = 0
        metric['recall_rcnn_%s' % str(cur_thresh)] = 0

    dataset = dataloader.dataset
    class_names = dataset.class_names
    det_annos = []

    if getattr(args, 'infer_time', False):
        start_iter = int(len(dataloader) * 0.1)
        infer_time_meter = common_utils.AverageMeter()

    logger.info('*************** EPOCH %s EVALUATION *****************' % epoch_id)
    if dist_test:
        num_gpus = torch.cuda.device_count()
        local_rank = cfg.LOCAL_RANK % num_gpus
        model = torch.nn.parallel.DistributedDataParallel(
                model,
                device_ids=[local_rank],
                broadcast_buffers=False
        )
    model.eval()

    if cfg.LOCAL_RANK == 0:
        progress_bar = tqdm.tqdm(total=len(dataloader), leave=True, desc='eval', dynamic_ncols=True)
    start_time = time.time()
    for i, batch_dict in enumerate(dataloader):
        load_data_to_gpu(batch_dict)

        if getattr(args, 'infer_time', False):
            start_time = time.time()

        with torch.no_grad():
            pred_dicts, ret_dict = model(batch_dict)

        disp_dict = {}

        if getattr(args, 'infer_time', False):
            inference_time = time.time() - start_time
            infer_time_meter.update(inference_time * 1000)
            # use ms to measure inference time
            disp_dict['infer_time'] = f'{infer_time_meter.val:.2f}({infer_time_meter.avg:.2f})'

        statistics_info(cfg, ret_dict, metric, disp_dict)
        annos = dataset.generate_prediction_dicts(
            batch_dict, pred_dicts, class_names,
            output_path=final_output_dir if args.save_to_file else None
        )
        det_annos += annos
        if cfg.LOCAL_RANK == 0:
            progress_bar.set_postfix(disp_dict)
            progress_bar.update()

    # file = 'result_center_partition_0.6.txt'
    # with open(file, 'w') as f: 
    #     json.dump(count_result, f)

    # file_gt = 'waymo_gt_list.txt'
    # with open(file_gt, 'w') as f: 
    #     json.dump(gt_list, f)

    if cfg.LOCAL_RANK == 0:
        progress_bar.close()

    if dist_test:
        rank, world_size = common_utils.get_dist_info()
        det_annos = common_utils.merge_results_dist(det_annos, len(dataset), tmpdir=result_dir / 'tmpdir')
        metric = common_utils.merge_results_dist([metric], world_size, tmpdir=result_dir / 'tmpdir')

    logger.info('*************** Performance of EPOCH %s *****************' % epoch_id)
    sec_per_example = (time.time() - start_time) / len(dataloader.dataset)
    logger.info('Generate label finished(sec_per_example: %.4f second).' % sec_per_example)

    if cfg.LOCAL_RANK != 0:
        return {}

    ret_dict = {}
    if dist_test:
        for key, val in metric[0].items():
            for k in range(1, world_size):
                metric[0][key] += metric[k][key]
        metric = metric[0]

    gt_num_cnt = metric['gt_num']
    for cur_thresh in cfg.MODEL.POST_PROCESSING.RECALL_THRESH_LIST:
        cur_roi_recall = metric['recall_roi_%s' % str(cur_thresh)] / max(gt_num_cnt, 1)
        cur_rcnn_recall = metric['recall_rcnn_%s' % str(cur_thresh)] / max(gt_num_cnt, 1)
        logger.info('recall_roi_%s: %f' % (cur_thresh, cur_roi_recall))
        logger.info('recall_rcnn_%s: %f' % (cur_thresh, cur_rcnn_recall))
        ret_dict['recall/roi_%s' % str(cur_thresh)] = cur_roi_recall
        ret_dict['recall/rcnn_%s' % str(cur_thresh)] = cur_rcnn_recall

    total_pred_objects = 0
    for anno in det_annos:
        total_pred_objects += anno['name'].__len__()
    logger.info('Average predicted number of objects(%d samples): %.3f'
                % (len(det_annos), total_pred_objects / max(1, len(det_annos))))

    with open(result_dir / 'result.pkl', 'wb') as f:
        pickle.dump(det_annos, f)
    # end_time = time.time()
    # execution_time = end_time - start_time  # Compute elapsed time
    # print(f"Execution time: {execution_time:.6f} seconds")

    result_str, result_dict = dataset.evaluation(
        det_annos, class_names,
        eval_metric=cfg.MODEL.POST_PROCESSING.EVAL_METRIC,
        output_path=final_output_dir
    )

    print(result_dict)

    


    logger.info(result_str)
    ret_dict.update(result_dict)

    logger.info('Result is saved to %s' % result_dir)
    logger.info('****************Evaluation done.*****************')
    
        # write_to_file_kitti(result_dict)
    write_to_file_nuscene(result_dict)
    # write_to_file_waymo(result_dict)
    
    return ret_dict


def write_to_file_nuscene(result_dict, model_name = 'centerpoint'):
    file_name = 'nuscene_result.xlsx'
    if not os.path.isfile(file_name):
        wb = Workbook()
    else:
        wb = load_workbook('nuscene_result.xlsx')

    if model_name in wb.sheetnames:
        sheet = wb[model_name]
    else:
        sheet = wb.create_sheet(model_name)
        sheet.title = model_name
    sheet.cell(row=1, column=column_index_from_string('C')).value = 'car'
    sheet.cell(row=1, column=column_index_from_string('D')).value = 'truck'
    sheet.cell(row=1, column=column_index_from_string('E')).value = 'construction_vehicle'
    sheet.cell(row=1, column=column_index_from_string('F')).value = 'bus'
    sheet.cell(row=1, column=column_index_from_string('G')).value = 'trailer'
    sheet.cell(row=1, column=column_index_from_string('H')).value = 'barrier'
    sheet.cell(row=1, column=column_index_from_string('I')).value = 'motorcycle'
    sheet.cell(row=1, column=column_index_from_string('J')).value = 'bicycle'
    sheet.cell(row=1, column=column_index_from_string('K')).value = 'pedestrian'
    sheet.cell(row=1, column=column_index_from_string('L')).value = 'traffic_cone'
    sheet.cell(row=1, column=column_index_from_string('M')).value = 'Overall'

    row = sheet.max_row +1
    print('row_number: ', row)

    sheet.cell(row=row, column=column_index_from_string('A')).value = sample_method
    sheet.cell(row=row, column=column_index_from_string('B')).value = sample_rate

    sheet.cell(row=row, column=column_index_from_string('C')).value = result_dict['cate']['car']['mean_ap']
    sheet.cell(row=row, column=column_index_from_string('D')).value = result_dict['cate']['truck']['mean_ap']
    sheet.cell(row=row, column=column_index_from_string('E')).value = result_dict['cate']['construction_vehicle']['mean_ap']

    sheet.cell(row=row, column=column_index_from_string('F')).value = result_dict['cate']['bus']['mean_ap']
    sheet.cell(row=row, column=column_index_from_string('G')).value = result_dict['cate']['trailer']['mean_ap']
    sheet.cell(row=row, column=column_index_from_string('H')).value = result_dict['cate']['barrier']['mean_ap']

    sheet.cell(row=row, column=column_index_from_string('I')).value = result_dict['cate']['motorcycle']['mean_ap']
    sheet.cell(row=row, column=column_index_from_string('J')).value = result_dict['cate']['bicycle']['mean_ap']
    sheet.cell(row=row, column=column_index_from_string('K')).value = result_dict['cate']['pedestrian']['mean_ap']
    sheet.cell(row=row, column=column_index_from_string('L')).value = result_dict['cate']['traffic_cone']['mean_ap']

    sheet.cell(row=row, column=column_index_from_string('M')).value = result_dict['mAP']

    wb.save(file_name)

def write_to_file_waymo(result_dict, model_name = 'centerpoint'):
    file_name = 'waymo_result.xlsx'
    if not os.path.isfile(file_name):
        wb = Workbook()
    else:
        wb = load_workbook('waymo_result.xlsx')

    if model_name in wb.sheetnames:
        sheet = wb[model_name]
    else:
        sheet = wb.create_sheet(model_name)
        sheet.title = model_name
    sheet.cell(row=1, column=column_index_from_string('C')).value = 'VEHICLE_L1'
    sheet.cell(row=1, column=column_index_from_string('D')).value = 'VEHICLE_L2'
    sheet.cell(row=1, column=column_index_from_string('E')).value = 'PEDESTRIAN_L1'
    sheet.cell(row=1, column=column_index_from_string('F')).value = 'PEDESTRIAN_L2'
    # sheet.cell(row=1, column=column_index_from_string('G')).value = 'SIGN_L1'
    # sheet.cell(row=1, column=column_index_from_string('H')).value = 'SIGN_L2'
    sheet.cell(row=1, column=column_index_from_string('G')).value = 'CYCLIST_L1'
    sheet.cell(row=1, column=column_index_from_string('H')).value = 'CYCLIST_L2'


    row = sheet.max_row +1
    print('row_number: ', row)

    sheet.cell(row=row, column=column_index_from_string('A')).value = sample_method
    sheet.cell(row=row, column=column_index_from_string('B')).value = sample_rate

    sheet.cell(row=row, column=column_index_from_string('C')).value = result_dict['OBJECT_TYPE_TYPE_VEHICLE_LEVEL_1/AP']
    sheet.cell(row=row, column=column_index_from_string('D')).value = result_dict['OBJECT_TYPE_TYPE_VEHICLE_LEVEL_2/AP']
    sheet.cell(row=row, column=column_index_from_string('E')).value = result_dict['OBJECT_TYPE_TYPE_PEDESTRIAN_LEVEL_1/AP']
    sheet.cell(row=row, column=column_index_from_string('F')).value = result_dict['OBJECT_TYPE_TYPE_PEDESTRIAN_LEVEL_2/AP']
    # sheet.cell(row=row, column=column_index_from_string('G')).value = result_dict['OBJECT_TYPE_TYPE_SIGN_LEVEL_1/AP']
    # sheet.cell(row=row, column=column_index_from_string('H')).value = result_dict['OBJECT_TYPE_TYPE_SIGN_LEVEL_2/AP']
    sheet.cell(row=row, column=column_index_from_string('G')).value = result_dict['OBJECT_TYPE_TYPE_CYCLIST_LEVEL_1/AP']
    sheet.cell(row=row, column=column_index_from_string('H')).value = result_dict['OBJECT_TYPE_TYPE_CYCLIST_LEVEL_2/AP']

    # sheet.cell(row=row, column=column_index_from_string('K')).value = result_dict['cate']['pedestrian']['mean_ap']
    # sheet.cell(row=row, column=column_index_from_string('L')).value = result_dict['cate']['traffic_cone']['mean_ap']

    # sheet.cell(row=row, column=column_index_from_string('M')).value = result_dict['mAP']

    wb.save(file_name)

def write_to_file_kitti(result_dict, model_name = 'centerpoint'):
    file_name = 'kitti_result.xlsx'
    if not os.path.isfile(file_name):
        wb = Workbook()
    else:
        wb = load_workbook('kitti_result.xlsx')

    if model_name in wb.sheetnames:
        sheet = wb[model_name]
    else:
        sheet = wb.create_sheet(model_name)
        sheet.title = model_name

    # sheet.cell(row=1, column=column_index_from_string('A')).value = 'Sample Method'
    # sheet.cell(row=1, column=column_index_from_string('B')).value = 'Sample Rate'

    sheet.cell(row=1, column=column_index_from_string('C')).value = 'Car'
    sheet.cell(row=1, column=column_index_from_string('F')).value = 'Pedestrian'
    sheet.cell(row=1, column=column_index_from_string('I')).value = 'Cyclist'

    sheet.cell(row=1, column=column_index_from_string('M')).value = 'Overall'


    row = sheet.max_row +1
    print('row_number: ', row)
    sheet.cell(row=row, column=column_index_from_string('A')).value = sample_method
    sheet.cell(row=row, column=column_index_from_string('B')).value = sample_rate

    sheet.cell(row=row, column=column_index_from_string('C')).value = result_dict['Car_3d/easy']
    sheet.cell(row=row, column=column_index_from_string('D')).value = result_dict['Car_3d/moderate']
    sheet.cell(row=row, column=column_index_from_string('E')).value = result_dict['Car_3d/hard']

    sheet.cell(row=row, column=column_index_from_string('F')).value = result_dict['Pedestrian_3d/easy']
    sheet.cell(row=row, column=column_index_from_string('G')).value = result_dict['Pedestrian_3d/moderate']
    sheet.cell(row=row, column=column_index_from_string('H')).value = result_dict['Pedestrian_3d/hard']

    sheet.cell(row=row, column=column_index_from_string('I')).value = result_dict['Cyclist_3d/easy']
    sheet.cell(row=row, column=column_index_from_string('J')).value = result_dict['Cyclist_3d/moderate']
    sheet.cell(row=row, column=column_index_from_string('K')).value = result_dict['Cyclist_3d/hard']

    
    # sheet.cell(row=row, column=column_index_from_string('O')).value = time

    # path = '/home/xiaoyu/experiments/mmdetection3d/tools/instance_sta.txt'
    # if os.path.exists(path):
    #     instance_point_dict = json.load(open(path))

    #     sheet.cell(row=row, column=column_index_from_string('P')).value = instance_point_dict['Car'][1]/instance_point_dict['Car'][0]
    #     sheet.cell(row=row, column=column_index_from_string('Q')).value = instance_point_dict['Pedestrian'][1]/instance_point_dict['Pedestrian'][0]
    #     sheet.cell(row=row, column=column_index_from_string('R')).value = instance_point_dict['Cyclist'][1]/instance_point_dict['Cyclist'][0]

    wb.save(file_name)

if __name__ == '__main__':
    pass
