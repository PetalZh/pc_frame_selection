import copy
import pickle
import random

import numpy as np
from skimage import io
from collections import Counter
import heapq, json

from . import kitti_utils
from ...ops.roiaware_pool3d import roiaware_pool3d_utils
from ...utils import box_utils, calibration_kitti, common_utils, object3d_kitti
from ..dataset import DatasetTemplate
import pcdet.constants as constants
import json, os
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import Workbook


class KittiDataset(DatasetTemplate):
    def __init__(self, dataset_cfg, class_names, training=True, root_path=None, logger=None):
        """
        Args:
            root_path:
            dataset_cfg:
            class_names:
            training:
            logger:
        """
        super().__init__(
            dataset_cfg=dataset_cfg, class_names=class_names, training=training, root_path=root_path, logger=logger
        )
        self.split = self.dataset_cfg.DATA_SPLIT[self.mode]
        self.root_split_path = self.root_path / ('training' if self.split != 'test' else 'testing')

        split_dir = self.root_path / 'ImageSets' / (self.split + '.txt')
        self.sample_id_list = [x.strip() for x in open(split_dir).readlines()] if split_dir.exists() else None

        self.kitti_infos = []
        self.include_kitti_data(self.mode)

    def getSceneCount(self, infos):
        object_classes = ['Car', 'Pedestrian', 'Cyclist']
        scene_counts = []

        for info in infos:
            obj_list = info['annos']['name']
            count_dict = Counter(obj_list)
            # Extract count for each class, default to 0 if missing
            counts = tuple(count_dict.get(cls, 0) for cls in object_classes)
            scene_counts.append(counts)
        return scene_counts
    def top_k_balanced_select(self, infos, M):
        scene_counts = self.getSceneCount(infos)
        k = M // 3
        selected_indices = set()

        for cls_idx in range(3):  # 0: Car, 1: Ped, 2: Cyclist
            top_k = heapq.nlargest(k, enumerate(scene_counts), key=lambda x: x[1][cls_idx])
            selected_indices.update(idx for idx, _ in top_k)

        # Fill in remaining if needed
        if len(selected_indices) < M:
            remaining = set(range(len(infos))) - selected_indices
            additional = list(remaining)[:M - len(selected_indices)]
            selected_indices.update(additional)

        return [infos[i] for i in list(selected_indices)[:M]]
    def greedy_select(self, infos, M):
        scene_counts = self.getSceneCount(infos)
        selected_indices = []
        total_counts = np.zeros(3)  # (cars, peds, cyclists)

        while len(selected_indices) < M:
            best_score = -1
            best_idx = None

            for idx, counts in enumerate(scene_counts):
                if idx in selected_indices:
                    continue
                # Score = contribution weighted by inverse of current totals
                score = sum(count / (1 + total) for count, total in zip(counts, total_counts))
                if score > best_score:
                    best_score = score
                    best_idx = idx

            if best_idx is not None:
                selected_indices.append(best_idx)
                total_counts += np.array(scene_counts[best_idx])
            else:
                break  # All remaining scores are 0

        return [infos[i] for i in selected_indices]
    
    def top_k_loss(self, infos, pre_trained_infos, k):
        # with open('output_loss_kitti_2.txt', 'r') as f:
        with open('output_loss_kitti_car_only.txt', 'r') as f:
            loss_dict = json.load(f)
            paired = list(zip(loss_dict['loss_list'], loss_dict['id_list']))
            paired_sorted = sorted(paired, key=lambda x: x[0], reverse=True)

            info_map = {info['point_cloud']['lidar_idx']: info for info in infos}
            pre_trained_set = {info['point_cloud']['lidar_idx']: info for info in pre_trained_infos}
            selected_infos  = []

            for _, idx in paired_sorted:
                if idx in pre_trained_set:          # already used → skip
                    continue
                if idx not in info_map:             # ID not found in `infos` → skip
                    continue

                selected_infos.append(info_map[idx])

                if len(selected_infos) == k:        # done
                    break
            return selected_infos
    def top_k_heatmap_emd(self, infos, pre_trained_infos, k):
        # with open('output_loss_kitti_2.txt', 'r') as f:
        #kitti_heatmap_EMD_1_10_10
        with open('kitti_heatmap_EMD_1_10_10.txt', 'r') as f:
            loss_dict = json.load(f)
            print('lenght of heatmap: {}'.format(len(loss_dict['score_list'])))
            paired = list(zip(loss_dict['score_list'], loss_dict['id_list']))
            paired_sorted = sorted(paired, key=lambda x: x[0], reverse=False)

            info_map = {info['point_cloud']['lidar_idx']: info for info in infos}
            pre_trained_set = {info['point_cloud']['lidar_idx']: info for info in pre_trained_infos}
            selected_infos  = []

            for _, idx in paired_sorted:
                if idx in pre_trained_set:          # already used → skip
                    continue
                if idx not in info_map:             # ID not found in `infos` → skip
                    continue

                selected_infos.append(info_map[idx])

                if len(selected_infos) == k:        # done
                    break
            # print('len(selected_infos): %d' % len(selected_infos))
            return selected_infos

    def top_k_loss_median(self, infos, pre_trained_infos, k):
        with open('output_loss_kitti_2.txt', 'r') as f:
            loss_dict = json.load(f)
            paired = list(zip(loss_dict['loss_list'], loss_dict['id_list']))
            # paired_sorted = sorted(paired, key=lambda x: x[0], reverse=True)

            median_loss = float(np.median([l for l, _ in paired]))
            
            candidates = [(l, idx) for l, idx in paired if l > median_loss]

            info_map = {info['point_cloud']['lidar_idx']: info for info in infos}
            pre_trained_ids = {info['point_cloud']['lidar_idx']: info for info in pre_trained_infos}
            valid = [(l, idx) for l, idx in candidates if idx in info_map and idx not in pre_trained_ids]

            if len(valid) < k:
                k = len(valid)
            idx_list = [idx for _, idx in valid]
            rng = np.random.default_rng(25)
            sampled_idxs = rng.choice(idx_list, size=k, replace=False)
            
            return [info_map[idx] for idx in sampled_idxs]
    
    def top_divergence_median(self, infos, pre_trained_infos, k):
        with open('output_divergence_kitti_2.txt', 'r') as f:
            loss_dict = json.load(f)
            paired = list(zip(loss_dict['loss_list'], loss_dict['id_list']))
            # paired_sorted = sorted(paired, key=lambda x: x[0], reverse=True)

            median_loss = float(np.median([l for l, _ in paired]))
            
            candidates = [(l, idx) for l, idx in paired if l > median_loss]

            info_map = {info['point_cloud']['lidar_idx']: info for info in infos}
            pre_trained_ids = {info['point_cloud']['lidar_idx']: info for info in pre_trained_infos}
            valid = [(l, idx) for l, idx in candidates if idx in info_map and idx not in pre_trained_ids]

            if len(valid) < k:
                k = len(valid)
            idx_list = [idx for _, idx in valid]
            rng = np.random.default_rng(25)
            sampled_idxs = rng.choice(idx_list, size=k, replace=False)
            
            return [info_map[idx] for idx in sampled_idxs]
    
    
    
    def top_k_divergence(self, infos, pre_trained_infos, k):
        # with open('output_divergence_kitti_2.txt', 'r') as f:
        with open('output_divergence_kitti_car_only.txt', 'r') as f:
            loss_dict = json.load(f)
            paired = list(zip(loss_dict['loss_list'], loss_dict['id_list']))
            paired_sorted = sorted(paired, key=lambda x: x[0], reverse=True)

            info_map = {info['point_cloud']['lidar_idx']: info for info in infos}
            pre_trained_set = {info['point_cloud']['lidar_idx']: info for info in pre_trained_infos}
            selected_infos  = []
            print(paired_sorted[:10])

            for _, idx in paired_sorted:
                if idx in pre_trained_set:          # already used → skip
                    continue
                if idx not in info_map:             # ID not found in `infos` → skip
                    continue

                selected_infos.append(info_map[idx])

                if len(selected_infos) == k:        # done
                    break
            return selected_infos
    
    def top_distance_euclidean(self, infos, k):
        with open('kitti_feature_distance.txt', 'r') as f:
            sorted_idx = json.load(f)

            info_map = {info['point_cloud']['lidar_idx']: info for info in infos}

            selected_infos = []
            for idx in sorted_idx:
                selected_infos.append(info_map[idx])
                if len(selected_infos) == k:        # done
                    break
            return selected_infos
        

    def random_select(self, infos, selected_infos, M):
        random.seed(25)
        selected_ids = {info['point_cloud']['lidar_idx'] for info in selected_infos}
        candidates = [info for info in infos if info['point_cloud']['lidar_idx'] not in selected_ids]
        
        selected_infos = random.sample(candidates, M)

        # print(f"Randomly selected {len(selected_infos)} items from index {start_idx} to {len(infos)}")

        return selected_infos

    def select_kitti_infos(self, infos, required_total=500):
        selected_infos = []

        # Priority buckets
        priority_1 = []  # Contains Car, Pedestrian, and Cyclist
        priority_2 = []  # Contains Pedestrian and Cyclist
        priority_3 = []  # Contains at least one of Cyclist, Pedestrian, or Car

        for info in infos:
            obj_set = set(info['annos']['name'])

            if {'Car', 'Pedestrian', 'Cyclist'} <= obj_set:
                priority_1.append(info)
            elif {'Pedestrian', 'Cyclist'} <= obj_set:
                priority_2.append(info)
            elif 'Cyclist' in obj_set or 'Pedestrian' in obj_set or 'Car' in obj_set:
                priority_3.append(info)

        def take_up_to(source, limit, taken):
            remaining = limit - len(taken)
            return taken + source[:remaining]

        selected_infos = take_up_to(priority_1, required_total, selected_infos)
        selected_infos = take_up_to(priority_2, required_total, selected_infos)
        selected_infos = take_up_to(priority_3, required_total, selected_infos)

        print(f"Selected {len(selected_infos)} items:")
        print(f"  From Priority 1: {min(len(priority_1), required_total)}")
        print(f"  From Priority 2: {min(len(priority_2), max(0, required_total - len(priority_1)))}")
        print(f"  From Priority 3: {max(0, required_total - len(priority_1) - len(priority_2))}")

        return selected_infos

    
    def getPretrain(self, infos, required_total=500):
        # infos.sort(key=lambda x: x['point_cloud']['lidar_idx'])
        # return infos[:required_total]
        result = []
        for info in infos:
            object_names = info['annos']['name'] 
            # if all(name == 'Car' for name in object_names):
            #     result.append(info)
            if all(name != 'pedestrian' and name != 'cyclist' for name in object_names):
                result.append(info)
            if len(result) == required_total:
                break
        return result


    def getSampledData(self, infos):
        sample_rate = constants.sample_rate
        sample_method = constants.sample_method

        # pretrain_size = int(len(infos)* 0.5)
        # kitti_infos = self.getPretrain(infos, required_total=pretrain_size)
            
        kitti_infos = self.getPretrain(infos, 500) #self.getPretrain(infos, 500) #infos[:500] 
        sample_size = int((len(infos)) * sample_rate)
        print('length of full: {}'.format(len(infos)))
        print('sample_size: %d' % sample_size)

        # for info in infos[:50]:
        #     print('info annos: ', info['annos']['name'])
        #     print('info annos: ', info['annos']['difficulty'])
            # break
        
        if sample_method == 'random':
            kitti_infos.extend(self.random_select(infos, kitti_infos, sample_size))
        elif sample_method == 'loss':
            kitti_infos.extend(self.top_k_loss(infos, kitti_infos, sample_size))
        elif sample_method == 'euclidean':
            kitti_infos.extend(self.top_distance_euclidean(infos, sample_size))
        elif sample_method == 'divergence':
            kitti_infos.extend(self.top_k_divergence(infos, kitti_infos, sample_size))
        elif sample_method == 'loss_median':
            kitti_infos.extend(self.top_k_loss_median(infos, infos[:500], sample_size))
        elif sample_method == 'divergence_median':
            kitti_infos.extend(self.top_divergence_median(infos, infos[:500], sample_size))
        elif sample_method == 'heatmap_emd':
            kitti_infos.extend(self.top_k_heatmap_emd(infos, infos[:500], sample_size))
        elif sample_method == 'pretrain':
            kitti_infos.extend(infos[:500])
        elif sample_method == 'full':
            kitti_infos = infos
        return kitti_infos

    def include_kitti_data(self, mode):
        if self.logger is not None:
            self.logger.info('Loading KITTI dataset')
        kitti_infos = []

        for info_path in self.dataset_cfg.INFO_PATH[mode]:
            info_path = self.root_path / info_path

            if not info_path.exists():
                continue
            with open(info_path, 'rb') as f:
                infos = pickle.load(f)

                if constants.isTrain:
                    kitti_infos = self.getSampledData(infos)
                    self.getSta(kitti_infos)
                    
                else:
                    kitti_infos.extend(infos)
                print('kitti_infos len: %d' % len(kitti_infos))

        self.kitti_infos.extend(kitti_infos)
        

        if self.logger is not None:
            self.logger.info('Total samples for KITTI dataset: %d' % (len(kitti_infos)))

    # def getSta(self, kitti_infos):
    #     sta_dict = {"Car": 0, "Pedestrian": 0, "Cyclist": 0}
    #     for info in kitti_infos:
    #         objs = info['annos']['name']
    #         if 'Car' in objs:
    #             # get the number of cars in objs
    #             car_count = np.sum(objs =='Car')
    #             sta_dict['Car'] += car_count
    #         if 'Pedestrian' in objs:
    #             # get the number of pedestrians in objs
    #             pedestrian_count = np.sum(objs =='Pedestrian')
    #             sta_dict['Pedestrian'] += pedestrian_count
    #         if 'Cyclist' in objs:
    #             # get the number of cyclists in objs
    #             cyclist_count = np.sum(objs =='Cyclist')
    #             sta_dict['Cyclist'] += cyclist_count
    #     print('sta_dict: %s' % sta_dict)
    #     # self.write_to_file(sta_dict, constants.sample_method, constants.sample_rate)   #self.dataset_cfg.MODEL_NAME

    def getSta(self, kitti_infos):
        # initialize nested dict with counters
        sta_dict = {
            "Car": {"total": 0, "easy": 0, "mod": 0, "hard": 0},
            "Pedestrian": {"total": 0, "easy": 0, "mod": 0, "hard": 0},
            "Cyclist": {"total": 0, "easy": 0, "mod": 0, "hard": 0},
        }

        for info in kitti_infos:
            objs = info['annos']['name']
            diffs = info['annos']['difficulty']   # same length as objs

            for obj, diff in zip(objs, diffs):
                if obj not in sta_dict:
                    continue  # skip non-target classes

                # always count in total
                sta_dict[obj]["total"] += 1

                # only count valid difficulty splits
                if diff == 0:
                    sta_dict[obj]["easy"] += 1
                elif diff == 1:
                    sta_dict[obj]["mod"] += 1
                elif diff == 2:
                    sta_dict[obj]["hard"] += 1

        print('sta_dict: %s' % sta_dict)
        return sta_dict
    
    
    def write_to_file(self, sta_dict, sample_method = '', sample_rate = '', model_name = 'centerpoint'):
        file_name = 'kitti_sta.xlsx'
        if not os.path.isfile(file_name):
            wb = Workbook()
        else:
            wb = load_workbook('kitti_sta.xlsx')

        if model_name in wb.sheetnames:
            sheet = wb[model_name]
        else:
            sheet = wb.create_sheet(model_name)
            sheet.title = model_name

        sheet.cell(row=1, column=column_index_from_string('A')).value = 'Sample Method'
        sheet.cell(row=1, column=column_index_from_string('B')).value = 'Sample Rate'

        sheet.cell(row=1, column=column_index_from_string('C')).value = 'Sta'

        row = sheet.max_row +1
        print('row_number: ', row)
        sheet.cell(row=row, column=column_index_from_string('A')).value = sample_method
        sheet.cell(row=row, column=column_index_from_string('B')).value = sample_rate

        sheet.cell(row=row, column=column_index_from_string('C')).value = sta_dict["Car"]
        sheet.cell(row=row, column=column_index_from_string('D')).value = sta_dict["Pedestrian"]
        sheet.cell(row=row, column=column_index_from_string('E')).value = sta_dict["Cyclist"]

        wb.save(file_name)
    
    def set_split(self, split):
        super().__init__(
            dataset_cfg=self.dataset_cfg, class_names=self.class_names, training=self.training, root_path=self.root_path, logger=self.logger
        )
        self.split = split
        self.root_split_path = self.root_path / ('training' if self.split != 'test' else 'testing')

        split_dir = self.root_path / 'ImageSets' / (self.split + '.txt')
        self.sample_id_list = [x.strip() for x in open(split_dir).readlines()] if split_dir.exists() else None

    def get_lidar(self, idx):
        lidar_file = self.root_split_path / 'velodyne' / ('%s.bin' % idx)
        assert lidar_file.exists()
        return np.fromfile(str(lidar_file), dtype=np.float32).reshape(-1, 4)

    def get_image(self, idx):
        """
        Loads image for a sample
        Args:
            idx: int, Sample index
        Returns:
            image: (H, W, 3), RGB Image
        """
        img_file = self.root_split_path / 'image_2' / ('%s.png' % idx)
        assert img_file.exists()
        image = io.imread(img_file)
        image = image.astype(np.float32)
        image /= 255.0
        return image

    def get_image_shape(self, idx):
        img_file = self.root_split_path / 'image_2' / ('%s.png' % idx)
        print(img_file)
        assert img_file.exists()
        return np.array(io.imread(img_file).shape[:2], dtype=np.int32)

    def get_label(self, idx):
        label_file = self.root_split_path / 'label_2' / ('%s.txt' % idx)
        assert label_file.exists()
        return object3d_kitti.get_objects_from_label(label_file)

    def get_depth_map(self, idx):
        """
        Loads depth map for a sample
        Args:
            idx: str, Sample index
        Returns:
            depth: (H, W), Depth map
        """
        depth_file = self.root_split_path / 'depth_2' / ('%s.png' % idx)
        assert depth_file.exists()
        depth = io.imread(depth_file)
        depth = depth.astype(np.float32)
        depth /= 256.0
        return depth

    def get_calib(self, idx):
        calib_file = self.root_split_path / 'calib' / ('%s.txt' % idx)
        assert calib_file.exists()
        return calibration_kitti.Calibration(calib_file)

    def get_road_plane(self, idx):
        plane_file = self.root_split_path / 'planes' / ('%s.txt' % idx)
        if not plane_file.exists():
            return None

        with open(plane_file, 'r') as f:
            lines = f.readlines()
        lines = [float(i) for i in lines[3].split()]
        plane = np.asarray(lines)

        # Ensure normal is always facing up, this is in the rectified camera coordinate
        if plane[1] > 0:
            plane = -plane

        norm = np.linalg.norm(plane[0:3])
        plane = plane / norm
        return plane

    @staticmethod
    def get_fov_flag(pts_rect, img_shape, calib):
        """
        Args:
            pts_rect:
            img_shape:
            calib:

        Returns:

        """
        pts_img, pts_rect_depth = calib.rect_to_img(pts_rect)
        val_flag_1 = np.logical_and(pts_img[:, 0] >= 0, pts_img[:, 0] < img_shape[1])
        val_flag_2 = np.logical_and(pts_img[:, 1] >= 0, pts_img[:, 1] < img_shape[0])
        val_flag_merge = np.logical_and(val_flag_1, val_flag_2)
        pts_valid_flag = np.logical_and(val_flag_merge, pts_rect_depth >= 0)

        return pts_valid_flag

    def get_infos(self, num_workers=4, has_label=True, count_inside_pts=True, sample_id_list=None):
        import concurrent.futures as futures

        def process_single_scene(sample_idx):
            print('%s sample_idx: %s' % (self.split, sample_idx))
            info = {}
            pc_info = {'num_features': 4, 'lidar_idx': sample_idx}
            info['point_cloud'] = pc_info

            image_info = {'image_idx': sample_idx, 'image_shape': self.get_image_shape(sample_idx)}
            info['image'] = image_info
            calib = self.get_calib(sample_idx)

            P2 = np.concatenate([calib.P2, np.array([[0., 0., 0., 1.]])], axis=0)
            R0_4x4 = np.zeros([4, 4], dtype=calib.R0.dtype)
            R0_4x4[3, 3] = 1.
            R0_4x4[:3, :3] = calib.R0
            V2C_4x4 = np.concatenate([calib.V2C, np.array([[0., 0., 0., 1.]])], axis=0)
            calib_info = {'P2': P2, 'R0_rect': R0_4x4, 'Tr_velo_to_cam': V2C_4x4}

            info['calib'] = calib_info

            if has_label:
                obj_list = self.get_label(sample_idx)
                annotations = {}
                annotations['name'] = np.array([obj.cls_type for obj in obj_list])
                annotations['truncated'] = np.array([obj.truncation for obj in obj_list])
                annotations['occluded'] = np.array([obj.occlusion for obj in obj_list])
                annotations['alpha'] = np.array([obj.alpha for obj in obj_list])
                annotations['bbox'] = np.concatenate([obj.box2d.reshape(1, 4) for obj in obj_list], axis=0)
                annotations['dimensions'] = np.array([[obj.l, obj.h, obj.w] for obj in obj_list])  # lhw(camera) format
                annotations['location'] = np.concatenate([obj.loc.reshape(1, 3) for obj in obj_list], axis=0)
                annotations['rotation_y'] = np.array([obj.ry for obj in obj_list])
                annotations['score'] = np.array([obj.score for obj in obj_list])
                annotations['difficulty'] = np.array([obj.level for obj in obj_list], np.int32)

                num_objects = len([obj.cls_type for obj in obj_list if obj.cls_type != 'DontCare'])
                num_gt = len(annotations['name'])
                index = list(range(num_objects)) + [-1] * (num_gt - num_objects)
                annotations['index'] = np.array(index, dtype=np.int32)

                loc = annotations['location'][:num_objects]
                dims = annotations['dimensions'][:num_objects]
                rots = annotations['rotation_y'][:num_objects]
                loc_lidar = calib.rect_to_lidar(loc)
                l, h, w = dims[:, 0:1], dims[:, 1:2], dims[:, 2:3]
                loc_lidar[:, 2] += h[:, 0] / 2
                gt_boxes_lidar = np.concatenate([loc_lidar, l, w, h, -(np.pi / 2 + rots[..., np.newaxis])], axis=1)
                annotations['gt_boxes_lidar'] = gt_boxes_lidar

                info['annos'] = annotations

                if count_inside_pts:
                    points = self.get_lidar(sample_idx)
                    calib = self.get_calib(sample_idx)
                    pts_rect = calib.lidar_to_rect(points[:, 0:3])

                    fov_flag = self.get_fov_flag(pts_rect, info['image']['image_shape'], calib)
                    pts_fov = points[fov_flag]
                    corners_lidar = box_utils.boxes_to_corners_3d(gt_boxes_lidar)
                    num_points_in_gt = -np.ones(num_gt, dtype=np.int32)

                    for k in range(num_objects):
                        flag = box_utils.in_hull(pts_fov[:, 0:3], corners_lidar[k])
                        num_points_in_gt[k] = flag.sum()
                    annotations['num_points_in_gt'] = num_points_in_gt

            return info

        sample_id_list = sample_id_list if sample_id_list is not None else self.sample_id_list
        with futures.ThreadPoolExecutor(num_workers) as executor:
            infos = executor.map(process_single_scene, sample_id_list)
        return list(infos)

    def create_groundtruth_database(self, info_path=None, used_classes=None, split='train'):
        import torch

        database_save_path = Path(self.root_path) / ('gt_database' if split == 'train' else ('gt_database_%s' % split))
        db_info_save_path = Path(self.root_path) / ('kitti_dbinfos_%s.pkl' % split)

        database_save_path.mkdir(parents=True, exist_ok=True)
        all_db_infos = {}

        with open(info_path, 'rb') as f:
            infos = pickle.load(f)

        for k in range(len(infos)):
            print('gt_database sample: %d/%d' % (k + 1, len(infos)))
            info = infos[k]
            sample_idx = info['point_cloud']['lidar_idx']
            points = self.get_lidar(sample_idx)
            annos = info['annos']
            names = annos['name']
            difficulty = annos['difficulty']
            bbox = annos['bbox']
            gt_boxes = annos['gt_boxes_lidar']

            num_obj = gt_boxes.shape[0]
            point_indices = roiaware_pool3d_utils.points_in_boxes_cpu(
                torch.from_numpy(points[:, 0:3]), torch.from_numpy(gt_boxes)
            ).numpy()  # (nboxes, npoints)

            for i in range(num_obj):
                filename = '%s_%s_%d.bin' % (sample_idx, names[i], i)
                filepath = database_save_path / filename
                gt_points = points[point_indices[i] > 0]

                gt_points[:, :3] -= gt_boxes[i, :3]
                with open(filepath, 'w') as f:
                    gt_points.tofile(f)

                if (used_classes is None) or names[i] in used_classes:
                    db_path = str(filepath.relative_to(self.root_path))  # gt_database/xxxxx.bin
                    db_info = {'name': names[i], 'path': db_path, 'image_idx': sample_idx, 'gt_idx': i,
                               'box3d_lidar': gt_boxes[i], 'num_points_in_gt': gt_points.shape[0],
                               'difficulty': difficulty[i], 'bbox': bbox[i], 'score': annos['score'][i]}
                    if names[i] in all_db_infos:
                        all_db_infos[names[i]].append(db_info)
                    else:
                        all_db_infos[names[i]] = [db_info]
        for k, v in all_db_infos.items():
            print('Database %s: %d' % (k, len(v)))

        with open(db_info_save_path, 'wb') as f:
            pickle.dump(all_db_infos, f)

    @staticmethod
    def generate_prediction_dicts(batch_dict, pred_dicts, class_names, output_path=None):
        """
        Args:
            batch_dict:
                frame_id:
            pred_dicts: list of pred_dicts
                pred_boxes: (N, 7), Tensor
                pred_scores: (N), Tensor
                pred_labels: (N), Tensor
            class_names:
            output_path:

        Returns:

        """
        def get_template_prediction(num_samples):
            ret_dict = {
                'name': np.zeros(num_samples), 'truncated': np.zeros(num_samples),
                'occluded': np.zeros(num_samples), 'alpha': np.zeros(num_samples),
                'bbox': np.zeros([num_samples, 4]), 'dimensions': np.zeros([num_samples, 3]),
                'location': np.zeros([num_samples, 3]), 'rotation_y': np.zeros(num_samples),
                'score': np.zeros(num_samples), 'boxes_lidar': np.zeros([num_samples, 7])
            }
            return ret_dict

        def generate_single_sample_dict(batch_index, box_dict):
            pred_scores = box_dict['pred_scores'].cpu().numpy()
            pred_boxes = box_dict['pred_boxes'].cpu().numpy()
            pred_labels = box_dict['pred_labels'].cpu().numpy()
            pred_dict = get_template_prediction(pred_scores.shape[0])
            if pred_scores.shape[0] == 0:
                return pred_dict

            calib = batch_dict['calib'][batch_index]
            image_shape = batch_dict['image_shape'][batch_index].cpu().numpy()
            pred_boxes_camera = box_utils.boxes3d_lidar_to_kitti_camera(pred_boxes, calib)
            pred_boxes_img = box_utils.boxes3d_kitti_camera_to_imageboxes(
                pred_boxes_camera, calib, image_shape=image_shape
            )

            pred_dict['name'] = np.array(class_names)[pred_labels - 1]
            pred_dict['alpha'] = -np.arctan2(-pred_boxes[:, 1], pred_boxes[:, 0]) + pred_boxes_camera[:, 6]
            pred_dict['bbox'] = pred_boxes_img
            pred_dict['dimensions'] = pred_boxes_camera[:, 3:6]
            pred_dict['location'] = pred_boxes_camera[:, 0:3]
            pred_dict['rotation_y'] = pred_boxes_camera[:, 6]
            pred_dict['score'] = pred_scores
            pred_dict['boxes_lidar'] = pred_boxes

            return pred_dict

        annos = []
        for index, box_dict in enumerate(pred_dicts):
            frame_id = batch_dict['frame_id'][index]

            single_pred_dict = generate_single_sample_dict(index, box_dict)
            single_pred_dict['frame_id'] = frame_id
            annos.append(single_pred_dict)

            if output_path is not None:
                cur_det_file = output_path / ('%s.txt' % frame_id)
                with open(cur_det_file, 'w') as f:
                    bbox = single_pred_dict['bbox']
                    loc = single_pred_dict['location']
                    dims = single_pred_dict['dimensions']  # lhw -> hwl

                    for idx in range(len(bbox)):
                        print('%s -1 -1 %.4f %.4f %.4f %.4f %.4f %.4f %.4f %.4f %.4f %.4f %.4f %.4f %.4f'
                              % (single_pred_dict['name'][idx], single_pred_dict['alpha'][idx],
                                 bbox[idx][0], bbox[idx][1], bbox[idx][2], bbox[idx][3],
                                 dims[idx][1], dims[idx][2], dims[idx][0], loc[idx][0],
                                 loc[idx][1], loc[idx][2], single_pred_dict['rotation_y'][idx],
                                 single_pred_dict['score'][idx]), file=f)

        return annos

    def evaluation(self, det_annos, class_names, **kwargs):
        if 'annos' not in self.kitti_infos[0].keys():
            return None, {}

        from .kitti_object_eval_python import eval as kitti_eval

        eval_det_annos = copy.deepcopy(det_annos)
        eval_gt_annos = [copy.deepcopy(info['annos']) for info in self.kitti_infos]
        ap_result_str, ap_dict = kitti_eval.get_official_eval_result(eval_gt_annos, eval_det_annos, class_names)

        return ap_result_str, ap_dict

    def __len__(self):
        if self._merge_all_iters_to_one_epoch:
            return len(self.kitti_infos) * self.total_epochs

        return len(self.kitti_infos)

    def __getitem__(self, index):
        # index = 4
        if self._merge_all_iters_to_one_epoch:
            index = index % len(self.kitti_infos)

        info = copy.deepcopy(self.kitti_infos[index])

        sample_idx = info['point_cloud']['lidar_idx']
        img_shape = info['image']['image_shape']
        calib = self.get_calib(sample_idx)
        get_item_list = self.dataset_cfg.get('GET_ITEM_LIST', ['points'])

        input_dict = {
            'frame_id': sample_idx,
            'calib': calib,
        }

        if 'annos' in info:
            annos = info['annos']
            annos = common_utils.drop_info_with_name(annos, name='DontCare')
            loc, dims, rots = annos['location'], annos['dimensions'], annos['rotation_y']
            gt_names = annos['name']
            gt_boxes_camera = np.concatenate([loc, dims, rots[..., np.newaxis]], axis=1).astype(np.float32)
            gt_boxes_lidar = box_utils.boxes3d_kitti_camera_to_lidar(gt_boxes_camera, calib)

            input_dict.update({
                'gt_names': gt_names,
                'gt_boxes': gt_boxes_lidar
            })
            if "gt_boxes2d" in get_item_list:
                input_dict['gt_boxes2d'] = annos["bbox"]

            road_plane = self.get_road_plane(sample_idx)
            if road_plane is not None:
                input_dict['road_plane'] = road_plane

        if "points" in get_item_list:
            points = self.get_lidar(sample_idx)
            if self.dataset_cfg.FOV_POINTS_ONLY:
                pts_rect = calib.lidar_to_rect(points[:, 0:3])
                fov_flag = self.get_fov_flag(pts_rect, img_shape, calib)
                points = points[fov_flag]
            input_dict['points'] = points

        if "images" in get_item_list:
            input_dict['images'] = self.get_image(sample_idx)

        if "depth_maps" in get_item_list:
            input_dict['depth_maps'] = self.get_depth_map(sample_idx)

        if "calib_matricies" in get_item_list:
            input_dict["trans_lidar_to_cam"], input_dict["trans_cam_to_img"] = kitti_utils.calib_to_matricies(calib)

        input_dict['calib'] = calib
        data_dict = self.prepare_data(data_dict=input_dict)

        data_dict['image_shape'] = img_shape
        return data_dict


def create_kitti_infos(dataset_cfg, class_names, data_path, save_path, workers=4):
    dataset = KittiDataset(dataset_cfg=dataset_cfg, class_names=class_names, root_path=data_path, training=False)
    train_split, val_split = 'train', 'val'

    train_filename = save_path / ('kitti_infos_%s.pkl' % train_split)
    val_filename = save_path / ('kitti_infos_%s.pkl' % val_split)
    trainval_filename = save_path / 'kitti_infos_trainval.pkl'
    test_filename = save_path / 'kitti_infos_test.pkl'

    print('---------------Start to generate data infos---------------')

    dataset.set_split(train_split)
    kitti_infos_train = dataset.get_infos(num_workers=workers, has_label=True, count_inside_pts=True)
    with open(train_filename, 'wb') as f:
        pickle.dump(kitti_infos_train, f)
    print('Kitti info train file is saved to %s' % train_filename)

    dataset.set_split(val_split)
    kitti_infos_val = dataset.get_infos(num_workers=workers, has_label=True, count_inside_pts=True)
    with open(val_filename, 'wb') as f:
        pickle.dump(kitti_infos_val, f)
    print('Kitti info val file is saved to %s' % val_filename)

    with open(trainval_filename, 'wb') as f:
        pickle.dump(kitti_infos_train + kitti_infos_val, f)
    print('Kitti info trainval file is saved to %s' % trainval_filename)

    dataset.set_split('test')
    kitti_infos_test = dataset.get_infos(num_workers=workers, has_label=False, count_inside_pts=False)
    with open(test_filename, 'wb') as f:
        pickle.dump(kitti_infos_test, f)
    print('Kitti info test file is saved to %s' % test_filename)

    print('---------------Start create groundtruth database for data augmentation---------------')
    dataset.set_split(train_split)
    dataset.create_groundtruth_database(train_filename, split=train_split)

    print('---------------Data preparation Done---------------')


if __name__ == '__main__':
    import sys
    if sys.argv.__len__() > 1 and sys.argv[1] == 'create_kitti_infos':
        import yaml
        from pathlib import Path
        from easydict import EasyDict
        dataset_cfg = EasyDict(yaml.safe_load(open(sys.argv[2])))
        ROOT_DIR = Path('/data')#(Path(__file__).resolve().parent / '../../../').resolve()

        create_kitti_infos(
            dataset_cfg=dataset_cfg,
            class_names=['Car', 'Pedestrian', 'Cyclist'],
            data_path=ROOT_DIR / 'kitti_detection_3d',
            save_path=ROOT_DIR / 'kitti_detection_3d'
            # data_path=ROOT_DIR / 'data' / 'kitti',
            # save_path=ROOT_DIR / 'data' / 'kitti'
        )
